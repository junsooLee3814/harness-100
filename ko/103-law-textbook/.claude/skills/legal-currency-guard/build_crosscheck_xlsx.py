# -*- coding: utf-8 -*-
"""
원문대조_작업표.xlsx 생성기 (legal-currency-guard 스킬 자산)

세무사·변호사가 출판 전 원문 대조 작업을 할 때 손에 들고 쓰는 엑셀 작업표를
`02_verification_checklist.md`(주 데이터 소스) + `chapters/*.md`(마커 실위치)
+ `05_review/verify_*.md`(확인 완료 이력)에서 **기계적으로 재생성**한다.

원고가 바뀌어 마커 수가 달라져도 재실행만 하면 갱신된다.
사람이 엑셀에 기입한 상태·확인결과·확인자·확인일은 재생성 시 키(장+마커행)로 자동 승계한다.

사용법:
    python build_crosscheck_xlsx.py "_workspace/상속증여세법"
    python build_crosscheck_xlsx.py "_workspace/상속증여세법" --dry-run

원칙:
  - 체크리스트에 없는 항목을 창작하지 않는다. 모든 셀 값은 md 원문에서 온다.
  - 기존 xlsx의 사람 기입분은 덮어쓰기 전에 읽어 승계하고, 승계 실패분은 이력 시트에 남긴다.
"""

from __future__ import annotations

import glob
import io
import os
import re
import sys
from collections import OrderedDict, defaultdict

try:
    import openpyxl
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:  # pragma: no cover
    sys.stderr.write("openpyxl 이 필요합니다:  pip install openpyxl\n")
    raise

if hasattr(sys.stdout, "buffer"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

# ---------------------------------------------------------------------------
# 상수 — 서식 (기존 수동 제작본의 서식을 계승: 헤더 네이비 1F3864 + 흰 볼드, 틀고정)
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

SUB_FILL = PatternFill("solid", fgColor="D9E2F3")
SUB_FONT = Font(bold=True, size=10, color="1F3864")

TITLE_FONT = Font(bold=True, size=14, color="1F3864")
NOTE_FONT = Font(size=9, color="808080")

PRIORITY_FILL = {
    "🔴 최우선": PatternFill("solid", fgColor="FCE7E7"),
    "🟡 중요": PatternFill("solid", fgColor="FFF6DD"),
    "🟢 보완": PatternFill("solid", fgColor="EAF4E6"),
    "⚪ 표기": PatternFill("solid", fgColor="F0F0F0"),
}
PRIORITY_ORDER = ["🔴 최우선", "🟡 중요", "🟢 보완", "⚪ 표기"]
EMOJI_TO_PRIORITY = {"🔴": "🔴 최우선", "🟡": "🟡 중요", "🟢": "🟢 보완", "⚪": "⚪ 표기"}

STATUS_OPTIONS = ["미확인", "확인중", "확인완료", "수정필요", "해당없음"]

THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

SHEET_MAIN = "① 작업표"
SHEET_SUMMARY = "② 요약"
SHEET_HOWTO = "③ 확인 방법"
SHEET_DONE = "④ 확인 완료 이력"
SHEET_LEGACY = ["⑤ 조문 인용 색인", "⑥ 핵심 수치 30선"]
LEGACY_SOURCE = ["조문 인용 색인", "핵심 수치 30선"]

MARKER_RE = re.compile(r"\[검증필요(?:[^\[\]]|\[[^\[\]]*\])*\]")

LOG: list[str] = []


def log(msg: str = "") -> None:
    LOG.append(msg)
    print(msg)


# ---------------------------------------------------------------------------
# 체크리스트 md 파싱
# ---------------------------------------------------------------------------


def slice_section(lines: list[str], start_pat: str, end_pat: str) -> list[str]:
    """start_pat 로 시작하는 줄 다음부터 end_pat 로 시작하는 줄 직전까지."""
    s = e = None
    for i, ln in enumerate(lines):
        if s is None and re.match(start_pat, ln):
            s = i + 1
        elif s is not None and re.match(end_pat, ln):
            e = i
            break
    if s is None:
        return []
    return lines[s: e if e is not None else len(lines)]


def strip_md(text: str) -> str:
    """엑셀 셀에 넣기 위한 최소 정리 — 강조 기호만 제거하고 문장은 원문 유지."""
    t = text.strip()
    t = re.sub(r"\*\*(.+?)\*\*", r"\1", t)
    t = re.sub(r"(?<!\w)\*(?!\s)(.+?)(?<!\s)\*(?!\w)", r"\1", t)
    t = re.sub(r"`([^`]+)`", r"\1", t)
    t = re.sub(r"\[\[([^\]]+)\]\]", r"\1", t)
    return t.strip()


def chapter_sort_key(ch) -> tuple:
    return (99, 0) if ch == "부록" else (int(ch), 0)


def parse_chapter_table(lines: list[str]) -> "OrderedDict[object, dict]":
    """§0.3 장별 분포 표 → {장: {title, 마커, 🔴, 🟡, 🟢, ⚪}}"""
    out: "OrderedDict[object, dict]" = OrderedDict()
    for ln in lines:
        if not ln.strip().startswith("|"):
            continue
        cells = [strip_md(c) for c in ln.strip().strip("|").split("|")]
        if len(cells) < 7:
            continue
        ch = cells[0].strip()
        if ch in ("장", "", "합계") or set(ch) <= set("-: "):
            continue
        key = "부록" if ch == "부록" else (int(ch) if ch.isdigit() else None)
        if key is None:
            continue
        def as_int(v):
            v = v.strip()
            return int(v) if v.lstrip("-").isdigit() else 0
        out[key] = {
            "title": cells[1].strip(),
            "마커": as_int(cells[2]),
            "🔴": as_int(cells[3]),
            "🟡": as_int(cells[4]),
            "🟢": as_int(cells[5]),
            "⚪": as_int(cells[6]),
        }
    return out


ITEM_RE = re.compile(
    r"^-\s*\[( |x)\]\s*(🔴|🟡|🟢)?\s*\*\*(.+?)\*\*\s*"
    r"(?:\((L\d+)([^)]*)\))?\s*(?:—|-)?\s*(.*)$"
)


def parse_section4(lines: list[str]) -> tuple[list[dict], list[dict]]:
    """§4 장별 검증 항목 → (마커 항목 144, 보완 과제 2)"""
    items: list[dict] = []
    extras: list[dict] = []
    ch = None
    ch_title = ""
    for ln in lines:
        h = re.match(r"^###\s+제(\d+)장\s*·\s*(.+)$", ln)
        if h:
            ch, ch_title = int(h.group(1)), h.group(2).split(" (")[0].strip()
            continue
        h = re.match(r"^###\s+부록\s*(.*)$", ln)
        if h:
            ch, ch_title = "부록", "부록 A~K"
            continue
        if not ln.lstrip().startswith("- ["):
            continue
        m = ITEM_RE.match(ln.strip())
        if not m:
            continue
        checked, emoji, head, lref, lsuffix, body = m.groups()
        if not emoji:
            continue  # `- [x] **마커 없음.**` 같은 안내 줄
        rec = {
            "장": ch,
            "장제목": ch_title,
            "절위치": strip_md(head),
            "우선순위": EMOJI_TO_PRIORITY[emoji],
            "행": int(lref[1:]) if lref else None,
            "행부기": (lsuffix or "").strip(),
            "내용": strip_md(body),
            "완료표시": checked == "x",
        }
        (items if lref else extras).append(rec)
    return items, extras


NOTE_RE = re.compile(
    r"^-\s*⚪\s*(.+?)\s*\(L(\d+)(?:,\s*마커\s*(\d+)개)?\)\s*(?:—|-)\s*(.*)$"
)


def parse_section5(lines: list[str]) -> list[dict]:
    """§5 표기·상호참조 마커 → 10건 (한 줄에 마커 2개인 항목은 행 분해)"""
    out: list[dict] = []
    for ln in lines:
        m = NOTE_RE.match(ln.strip())
        if not m:
            continue
        where, lno, cnt, body = m.groups()
        where = strip_md(where)
        cm = re.match(r"^(\d+)장\s*(.*)$", where)
        if cm:
            ch, sec = int(cm.group(1)), cm.group(2).strip()
        elif where.startswith("부록"):
            ch, sec = "부록", where
        else:
            ch, sec = None, where
        for k in range(int(cnt or 1)):
            out.append({
                "장": ch,
                "장제목": "",
                "절위치": sec or where,
                "우선순위": "⚪ 표기",
                "행": int(lno),
                "행부기": f"마커 {k + 1}/{cnt}" if cnt else "",
                "내용": strip_md(body),
                "완료표시": False,
                "표기용": True,
            })
    return out


def parse_bullets(lines: list[str]) -> list[dict]:
    """`- [ ] 🔴 ...` / `- ...` 형태 불릿을 (우선순위, 텍스트, 완료) 로."""
    out = []
    for ln in lines:
        s = ln.strip()
        m = re.match(r"^-\s*\[( |x)\]\s*(🔴|🟡|🟢|⚪)?\s*(.*)$", s)
        if m:
            out.append({
                "완료": m.group(1) == "x",
                "우선순위": EMOJI_TO_PRIORITY.get(m.group(2) or "", ""),
                "text": strip_md(m.group(3)),
            })
            continue
        m = re.match(r"^\d+\.\s*(.+)$", s)
        if m:
            out.append({"완료": False, "우선순위": "", "text": strip_md(m.group(1))})
    return out


# ---------------------------------------------------------------------------
# 조문·출처 추출
# ---------------------------------------------------------------------------

LAW_WORDS = (
    "국세기본법|국세징수법|국제조세조정법|금융실명법|지방세법|소득세법|종부세법|종합부동산세법|"
    "상증세법|상증법|조특법|조특령|국조법|국기법|국기령|국기칙|민법|중소기업기본법|"
    "사무처리규정|시행규칙|시행령|훈령|법"
)
CITE_PATTERNS = [
    re.compile(r"(?:%s)\s*§\s*\d+(?:의\d+)?(?:\s*[①-⑳])*(?:\s*제?\d+호)?" % LAW_WORDS),
    re.compile(r"(?<![가-힣])§\s*\d+(?:의\d+)?(?:\s*[①-⑳])*(?:\s*제?\d+호)?"),
    re.compile(r"(?:%s)\s*제\d+조(?:의\d+)?(?:\s*제\d+항)?(?:\s*제\d+호)?" % LAW_WORDS),
    re.compile(r"사전-\d{4}-법규재산-\d{4}"),
    re.compile(r"서면-\d{4}-[가-힣]+-\d+"),
    re.compile(r"조심\s*\d{4}[가-힣]\d+"),
    re.compile(r"\d{4}[두누구다가나마](?:합)?\d+"),
    re.compile(r"\d{4}[가-힣]{0,2}합\d+"),
    re.compile(r"대법원\s*\d{4}\.\s*\d+\.\s*\d+\."),
    re.compile(r"훈령\s*제\d+호"),
    re.compile(r"법률\s*제\d+호"),
    re.compile(r"대통령령\s*제\d+호"),
]


def extract_cites(*texts: str) -> str:
    """겹치는 매치는 긴 쪽만 남긴다 ('시행령 §78①' vs '§78①' 중복 방지)."""
    seen: "OrderedDict[str, None]" = OrderedDict()
    for t in texts:
        if not t:
            continue
        spans = []
        for pat in CITE_PATTERNS:
            for m in pat.finditer(t):
                spans.append((m.start(), m.end(), m.group(0)))
        spans.sort(key=lambda s: (s[0], -(s[1] - s[0])))
        taken: list[tuple[int, int]] = []
        for st, en, raw in spans:
            if any(st < e and en > s for s, e in taken):
                continue
            taken.append((st, en))
            tok = re.sub(r"\s+", " ", raw).strip()
            if len(tok) >= 3:
                seen.setdefault(tok, None)
    return " · ".join(seen.keys())


def derive_status(text: str, priority: str) -> tuple[str, str]:
    """체크리스트 서술에서 상태와 비고를 도출한다 (창작 없음 — 표기 그대로 반영)."""
    verified = "✅ 검증 완료" in text
    partial = ("부분 확인" in text) or ("verify 유지 항목" in text)
    if verified and not partial and "유지(미확정)" not in text:
        return "확인완료", "verify_*.md 원문대조로 확정 — 남은 작업은 원고 마커 삭제 + 출처 각주"
    if verified:
        return "확인중", "일부만 확정(체크리스트에 ✅ + 유지 병기) — 잔여 쟁점 확인 필요"
    if partial:
        return "확인중", "verify로 근거 일부 확보, 항목은 유지"
    if priority == "⚪ 표기":
        return "해당없음", "표기 설명·상호참조 — 원문 조회 대상 아님. 마커를 지우지 말 것"
    return "미확인", ""


# ---------------------------------------------------------------------------
# 원고 마커 수집 + 체크리스트 항목과 짝짓기
# ---------------------------------------------------------------------------


def collect_manuscript_markers(chapters_dir: str) -> dict:
    """{장: [(파일명, 행번호, 마커원문), ...]} — 파일·행·등장순."""
    out = defaultdict(list)
    for path in sorted(glob.glob(os.path.join(chapters_dir, "*.md"))):
        base = os.path.basename(path)
        m = re.match(r"^(\d+)_", base)
        if not m:
            continue
        num = int(m.group(1))
        ch = "부록" if num >= 90 else num
        with open(path, encoding="utf-8") as fh:
            for i, line in enumerate(fh, 1):
                for mk in MARKER_RE.finditer(line):
                    out[ch].append((base, i, mk.group(0)))
    return out


def pair_with_manuscript(items: list[dict], markers: dict) -> tuple[int, int]:
    """장별로 행번호 오름차순 정렬 후 순서대로 1:1 짝짓기. (짝지음, 행번호 정확일치)"""
    paired = exact = 0
    by_ch = defaultdict(list)
    for it in items:
        by_ch[it["장"]].append(it)
    for ch, group in by_ch.items():
        group.sort(key=lambda r: (r["행"] or 0))
        found = sorted(markers.get(ch, []), key=lambda t: (t[1],))
        if len(group) != len(found):
            log(f"  [주의] 제{ch}장 — 체크리스트 {len(group)}건 vs 원고 마커 {len(found)}건 "
                f"(불일치: 원고 편집으로 마커가 증감했을 수 있음)")
        for idx, it in enumerate(group):
            if idx < len(found):
                fname, lno, text = found[idx]
                it["원고파일"] = fname
                it["원고행"] = lno
                it["마커원문"] = text
                paired += 1
                if lno == it["행"]:
                    exact += 1
            else:
                it["원고파일"] = ""
                it["원고행"] = None
                it["마커원문"] = ""
    return paired, exact


# ---------------------------------------------------------------------------
# 기존 파일 읽기 — 사람 기입분 승계
# ---------------------------------------------------------------------------


def read_existing(path: str) -> dict:
    """기존 xlsx에서 (1) 작업표 사람 기입분 (2) 승계 대상 레거시 시트 (3) 구본 기입 흔적"""
    info = {"entries": {}, "legacy": {}, "orphan_notes": [], "sheets": []}
    if not os.path.exists(path):
        log("  기존 파일 없음 — 신규 생성")
        return info
    wb = openpyxl.load_workbook(path)
    info["sheets"] = list(wb.sheetnames)
    log(f"  기존 시트: {info['sheets']}")

    if SHEET_MAIN in wb.sheetnames:
        ws = wb[SHEET_MAIN]
        hdr = {str(ws.cell(1, c).value).strip(): c for c in range(1, ws.max_column + 1)
               if ws.cell(1, c).value}
        kcol, acol, scol = hdr.get("키"), hdr.get("자동판정"), hdr.get("상태")
        if kcol and not acol:
            log("  [주의] 구 형식(자동판정 열 없음) — 상태 열은 스크립트 산출값과 구분할 수 "
                "없으므로 승계하지 않고, 사람 기입 열만 승계한다")
        if kcol:
            for r in range(2, ws.max_row + 1):
                key = ws.cell(r, kcol).value
                if not key:
                    continue
                vals = {}
                if acol and scol:
                    cur, auto = ws.cell(r, scol).value, ws.cell(r, acol).value
                    if cur not in (None, "") and cur != auto:
                        vals["상태"] = cur          # 사람이 바꾼 값만 승계
                for name in HUMAN_COLS:
                    c = hdr.get(name)
                    if c and ws.cell(r, c).value not in (None, ""):
                        vals[name] = ws.cell(r, c).value
                if vals:
                    info["entries"][str(key)] = vals
            log(f"  사람 기입분 승계 대상: {len(info['entries'])}행")

    # 구본(수동 제작본) 시트의 '기입' 열 흔적 — 승계 키가 없으므로 이력으로 남긴다
    for name, cols in (("검증필요", [3, 4, 5]), ("조문 인용 색인", [3, 4]),
                       ("핵심 수치 30선", [4, 5])):
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        hits = 0
        for r in range(2, ws.max_row + 1):
            for c in cols:
                v = ws.cell(r, c).value
                if v not in (None, ""):
                    hits += 1
                    info["orphan_notes"].append(
                        (name, r, str(ws.cell(1, c).value), str(v))
                    )
        log(f"  구본 '{name}' 기입 흔적: {hits}건")

    for src, dst in zip(LEGACY_SOURCE, SHEET_LEGACY):
        pick = src if src in wb.sheetnames else (dst if dst in wb.sheetnames else None)
        if not pick:
            continue
        ws = wb[pick]
        rows = [[ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
                for r in range(1, ws.max_row + 1)]
        widths = [ws.column_dimensions[get_column_letter(c)].width
                  for c in range(1, ws.max_column + 1)]
        info["legacy"][dst] = {"rows": rows, "widths": widths}
        log(f"  레거시 시트 승계: '{pick}' → '{dst}' ({len(rows)}행)")
    wb.close()
    return info


# ---------------------------------------------------------------------------
# 시트 작성 헬퍼
# ---------------------------------------------------------------------------


def write_header(ws, row: int, headers: list[str], widths: list[float] | None = None):
    for c, name in enumerate(headers, 1):
        cell = ws.cell(row, c, name)
        cell.fill, cell.font, cell.alignment, cell.border = (
            HEADER_FILL, HEADER_FONT, HEADER_ALIGN, BOX)
    if widths:
        for c, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[row].height = 30


def band(ws, row: int, text: str, span: int, fill=SUB_FILL, font=SUB_FONT):
    for c in range(1, span + 1):
        cell = ws.cell(row, c)
        cell.fill = fill
        if c == 1:
            cell.value = text
            cell.font = font
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    return row + 1


# ---------------------------------------------------------------------------
# ① 작업표
# ---------------------------------------------------------------------------

# 마지막 2열(키·자동판정)은 숨김 — 재생성 시 사람 기입분을 정확히 승계하기 위한 내부 열이다.
# '자동판정'은 생성 당시 스크립트가 계산한 기본 상태값. 재실행 시 상태 != 자동판정 이면
# 사람이 손댄 것으로 보고 보존하고, 같으면 새로 계산한 기본값이 이긴다.
MAIN_HEADERS = [
    "연번", "장", "절·위치", "우선순위", "확인 항목 (마커 원문)",
    "확인해야 할 조문·출처", "확인 내용 (체크리스트)", "원고 위치",
    "상태", "확인 결과 기재란", "근거 URL·법령버전", "확인자", "확인일", "비고",
    "키", "자동판정",
]
MAIN_WIDTHS = [5.5, 6, 28, 11, 58, 30, 74, 24, 10, 34, 28, 9, 12, 40, 11, 11]
HUMAN_COLS = ["확인 결과 기재란", "근거 URL·법령버전", "확인자", "확인일"]


def build_main(wb, items: list[dict], entries: dict) -> int:
    ws = wb.create_sheet(SHEET_MAIN)
    write_header(ws, 1, MAIN_HEADERS, MAIN_WIDTHS)

    restored = 0
    for n, it in enumerate(items, 1):
        r = n + 1
        key = "%s|L%04d|%d" % (it["장"], it["행"] or 0, it.get("중복", 1))
        status, note = derive_status(it["내용"] + " " + it.get("마커원문", ""),
                                     it["우선순위"])
        extra = []
        if it["행부기"]:
            extra.append(it["행부기"])
        if note:
            extra.append(note)
        for tag in re.findall(r"※[^※]+", it["내용"]):
            extra.append(strip_md(tag).rstrip("."))
        pos = (f'{it.get("원고파일", "")} L{it["원고행"]}'
               if it.get("원고행") else "(원고에서 마커 미발견)")

        vals = [
            n,
            it["장"],
            it["절위치"],
            it["우선순위"],
            it.get("마커원문", ""),
            extract_cites(it["내용"], it.get("마커원문", "")),
            it["내용"],
            pos,
            status,
            "", "", "", "",
            " / ".join(dict.fromkeys(extra)),
            key,
            status,
        ]
        saved = entries.get(key)
        if saved:
            restored += 1
            for name, v in saved.items():
                vals[MAIN_HEADERS.index(name)] = v

        fill = PRIORITY_FILL.get(it["우선순위"])
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r, c, v)
            cell.border = BOX
            cell.font = Font(size=10)
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=c in (3, 5, 6, 7, 8, 10, 11, 14),
                horizontal="center" if c in (1, 2, 4, 9, 12, 13) else "left",
            )
            if fill and c != 9:
                cell.fill = fill
        ws.cell(r, 13).number_format = "yyyy-mm-dd"

    last = len(items) + 1
    ws.freeze_panes = "E2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(MAIN_HEADERS) - 2)}{last}"
    for off in (1, 0):
        ws.column_dimensions[get_column_letter(len(MAIN_HEADERS) - off)].hidden = True

    dv = DataValidation(type="list", formula1='"%s"' % ",".join(STATUS_OPTIONS),
                        allow_blank=True, showDropDown=False)
    dv.error = "목록에서 고르세요: " + " / ".join(STATUS_OPTIONS)
    dv.errorTitle = "상태 값 확인"
    dv.prompt = "미확인 → 확인중 → 확인완료 / 원고 수정이 필요하면 수정필요"
    dv.promptTitle = "상태"
    ws.add_data_validation(dv)
    dv.add(f"I2:I{last}")

    rng = f"A2:N{last}"
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=['$I2="확인완료"'],
        fill=PatternFill("solid", fgColor="D6ECD2"), font=Font(color="5A7A55", size=10)))
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=['$I2="수정필요"'],
        fill=PatternFill("solid", fgColor="FBD5B5"), font=Font(bold=True, size=10)))

    if restored:
        log(f"  사람 기입분 승계 완료: {restored}행")
    return restored


# ---------------------------------------------------------------------------
# ② 요약
# ---------------------------------------------------------------------------


def build_summary(wb, items, chap_table, focus, common, extras, watch, counts):
    ws = wb.create_sheet(SHEET_SUMMARY)
    for c, w in enumerate([9, 34, 9, 8, 8, 8, 8, 11, 11, 60], 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    M = f"'{SHEET_MAIN}'"
    r = 1
    ws.cell(r, 1, "원문 대조 진행 현황").font = TITLE_FONT
    r += 1
    ws.cell(r, 1, "상태 열(① 작업표 I열)을 채우면 아래 수치·진행률이 자동으로 갱신된다.").font = NOTE_FONT
    r += 2

    r = band(ws, r, "■ 전체", 10)
    kpis = [
        ("총 등재 건수", f"=COUNTA({M}!$A$2:$A$2000)"),
        ("확인완료", f'=COUNTIF({M}!$I$2:$I$2000,"확인완료")'),
        ("확인중", f'=COUNTIF({M}!$I$2:$I$2000,"확인중")'),
        ("미확인", f'=COUNTIF({M}!$I$2:$I$2000,"미확인")'),
        ("수정필요", f'=COUNTIF({M}!$I$2:$I$2000,"수정필요")'),
        ("해당없음(표기용)", f'=COUNTIF({M}!$I$2:$I$2000,"해당없음")'),
    ]
    for name, f in kpis:
        ws.cell(r, 1, name).font = Font(bold=True, size=10)
        ws.cell(r, 3, f).alignment = Alignment(horizontal="center")
        r += 1
    ws.cell(r, 1, "전체 진행률").font = Font(bold=True, size=10)
    c = ws.cell(r, 3, f'=IFERROR(COUNTIF({M}!$I$2:$I$2000,"확인완료")/'
                      f'(COUNTA({M}!$A$2:$A$2000)-'
                      f'COUNTIF({M}!$I$2:$I$2000,"해당없음")),"")')
    c.number_format, c.alignment = "0.0%", Alignment(horizontal="center")
    ws.cell(r, 10, "분모는 ⚪ 표기용 10건을 뺀 실제 검증 대상 기준").font = NOTE_FONT
    r += 1
    ws.cell(r, 1, "🔴 최우선 진행률").font = Font(bold=True, size=10, color="C00000")
    c = ws.cell(r, 3, f'=IFERROR(COUNTIFS({M}!$D$2:$D$2000,"🔴 최우선",'
                      f'{M}!$I$2:$I$2000,"확인완료")/'
                      f'COUNTIF({M}!$D$2:$D$2000,"🔴 최우선"),"")')
    c.number_format, c.alignment = "0.0%", Alignment(horizontal="center")
    ws.cell(r, 10, "🔴은 어떤 경우에도 미해소 상태로 출판하지 말 것 (체크리스트 §0.2)").font = \
        Font(bold=True, size=10, color="C00000")
    r += 2

    r = band(ws, r, "■ 우선순위별", 10)
    write_header(ws, r, ["우선순위", "기준 — 틀렸을 때의 결과", "등재", "확인완료",
                         "확인중", "미확인", "수정필요", "해당없음", "진행률", ""])
    r += 1
    basis = {
        "🔴 최우선": "금액·세율·한도·비율·요건 연수·산식 — 세액이 틀어진다",
        "🟡 중요": "조문번호·항·호 배치, 법문 축자 표현, 준용 관계 — 인용이 틀린다",
        "🟢 보완": "예규·판례 사건번호, 서식 명칭, 절차·통계 — 신뢰도가 깎인다",
        "⚪ 표기": "표기 설명·상호참조 — 원문 조회 대상이 아니며 마커를 지우지 말 것",
    }
    for p in PRIORITY_ORDER:
        ws.cell(r, 1, p).alignment = Alignment(horizontal="center")
        ws.cell(r, 2, basis[p]).alignment = Alignment(wrap_text=True, vertical="center")
        ws.cell(r, 3, f'=COUNTIF({M}!$D$2:$D$2000,"{p}")')
        for off, st in enumerate(["확인완료", "확인중", "미확인", "수정필요", "해당없음"]):
            ws.cell(r, 4 + off,
                    f'=COUNTIFS({M}!$D$2:$D$2000,"{p}",{M}!$I$2:$I$2000,"{st}")')
        cc = ws.cell(r, 9, f'=IFERROR(D{r}/(C{r}-H{r}),"")')
        cc.number_format = "0.0%"
        for c in range(1, 10):
            ws.cell(r, c).border = BOX
            if c >= 3:
                ws.cell(r, c).alignment = Alignment(horizontal="center")
        ws.cell(r, 1).fill = PRIORITY_FILL[p]
        r += 1
    r += 1

    r = band(ws, r, "■ 장별", 10)
    write_header(ws, r, ["장", "제목(약칭)", "마커", "🔴", "🟡", "🟢", "⚪",
                         "확인완료", "진행률", "집중 검증 권고"])
    r += 1
    focus_map = {}
    for f in focus:
        for cm in re.findall(r"제(\d+)장", f["text"][:14]):
            focus_map[int(cm)] = f["text"]
    first_row = r
    for ch, meta in chap_table.items():
        crit = ch if isinstance(ch, int) else '"부록"'
        crit_s = str(ch) if isinstance(ch, int) else '"부록"'
        ws.cell(r, 1, ch).alignment = Alignment(horizontal="center")
        ws.cell(r, 2, meta["title"])
        ws.cell(r, 3, f"=COUNTIF({M}!$B$2:$B$2000,{crit_s})")
        for off, p in enumerate(PRIORITY_ORDER):
            ws.cell(r, 4 + off,
                    f'=COUNTIFS({M}!$B$2:$B$2000,{crit_s},{M}!$D$2:$D$2000,"{p}")')
        ws.cell(r, 8, f'=COUNTIFS({M}!$B$2:$B$2000,{crit_s},'
                      f'{M}!$I$2:$I$2000,"확인완료")')
        # 진행률 분모에서 ⚪ 표기용(G열)을 뺀다 — 상단 KPI 와 같은 기준
        cc = ws.cell(r, 9, f'=IFERROR(H{r}/(C{r}-G{r}),"")')
        cc.number_format = "0.0%"
        if isinstance(ch, int) and ch in focus_map:
            ws.cell(r, 10, focus_map[ch]).font = Font(bold=True, size=10, color="C00000")
            for c in range(1, 10):
                ws.cell(r, c).fill = PatternFill("solid", fgColor="FDECEC")
        ws.cell(r, 10).alignment = Alignment(wrap_text=True, vertical="center")
        for c in range(1, 11):
            ws.cell(r, c).border = BOX
            if 3 <= c <= 9:
                ws.cell(r, c).alignment = Alignment(horizontal="center")
        r += 1
    ws.cell(r, 1, "합계").font = Font(bold=True, size=10)
    for c in range(3, 9):
        cell = ws.cell(r, c, f"=SUM({get_column_letter(c)}{first_row}:"
                             f"{get_column_letter(c)}{r - 1})")
        cell.font = Font(bold=True, size=10)
        cell.alignment = Alignment(horizontal="center")
    cc = ws.cell(r, 9, f'=IFERROR(H{r}/(C{r}-G{r}),"")')
    cc.number_format, cc.font = "0.0%", Font(bold=True, size=10)
    for c in range(1, 10):
        ws.cell(r, c).border = BOX
        ws.cell(r, c).fill = PatternFill("solid", fgColor="D9E2F3")
    r += 2

    def block(row, title, rows, note=""):
        row = band(ws, row, title, 10)
        if note:
            ws.cell(row, 1, note).font = NOTE_FONT
            row += 1
        write_header(ws, row, ["상태", "우선순위", "내용", "", "", "", "", "", "", ""])
        ws.cell(row, 3).value = "내용"
        row += 1
        for b in rows:
            ws.cell(row, 1, "완료" if b.get("완료") else "미확인").alignment = \
                Alignment(horizontal="center")
            ws.cell(row, 2, b.get("우선순위", "")).alignment = \
                Alignment(horizontal="center")
            ws.cell(row, 3, b["text"]).alignment = Alignment(wrap_text=True, vertical="top")
            for c in range(1, 4):
                ws.cell(row, c).border = BOX
            if b.get("우선순위") in PRIORITY_FILL:
                ws.cell(row, 2).fill = PRIORITY_FILL[b["우선순위"]]
            row += 1
        return row + 1

    r = block(r, "■ 집중 검증 권고 장 — 여기부터 시작하라 (체크리스트 §0.4)", focus)
    r = block(r, "■ 전 장 공통 — 1회 스캔으로 끝내는 작업 (체크리스트 §0.5)", common)
    r = block(r, "■ 보완 과제 — 마커 외 (체크리스트 §4)",
              [{"완료": e["완료표시"], "우선순위": e["우선순위"],
                "text": f'{e["장"]}장 · {e["절위치"]} — {e["내용"]}'
                        if isinstance(e["장"], int) else
                        f'{e["장"]} · {e["절위치"]} — {e["내용"]}'} for e in extras])
    r = block(r, "■ 개정 감시 항목 — 매 세션·매 개정판 재확인 (체크리스트 §2)", watch,
              "출판 직전 이 블록을 다시 돌릴 것. 검증 시점과 출판 시점 사이에 개정이 끼면 검증 결과는 무효가 된다.")

    ws.freeze_panes = "A4"
    return ws


# ---------------------------------------------------------------------------
# ③ 확인 방법
# ---------------------------------------------------------------------------


def build_howto(wb, sec1_lines):
    ws = wb.create_sheet(SHEET_HOWTO)
    write_header(ws, 1, ["절", "구분", "내용", "", ""], [11, 20, 86, 26, 26])
    ws.cell(1, 3, "내용 / URL / 값")
    r = 2
    sec = ""
    in_code = False
    code_buf: list[str] = []
    table_hdr_done = False

    def put(kind, cells, mono=False, warn=False):
        nonlocal r
        ws.cell(r, 1, sec)
        ws.cell(r, 2, kind).font = Font(bold=True, size=10,
                                        color="C00000" if warn else "1F3864")
        for i, v in enumerate(cells):
            cell = ws.cell(r, 3 + i, v)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if mono:
                cell.font = Font(name="Consolas", size=10)
        for c in range(1, 6):
            ws.cell(r, c).border = BOX
            if c >= 3 and ws.cell(r, c).value is None:
                ws.cell(r, c).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(r, 1).alignment = Alignment(horizontal="center", vertical="top")
        ws.cell(r, 2).alignment = Alignment(vertical="top", wrap_text=True)
        r += 1

    for ln in sec1_lines:
        s = ln.rstrip()
        if s.strip().startswith("```"):
            if in_code:
                put("URL 패턴", ["\n".join(code_buf)], mono=True)
                code_buf = []
            in_code = not in_code
            continue
        if in_code:
            code_buf.append(s.strip())
            continue
        h = re.match(r"^###\s+(1\.\d+)\s*(.*)$", s)
        if h:
            sec = h.group(1)
            r = band(ws, r, f"{h.group(1)}  {strip_md(h.group(2))}", 5)
            table_hdr_done = False
            continue
        if not s.strip() or set(s.strip()) <= set("-*_ "):
            continue  # 빈 줄·수평선(---)은 옮기지 않는다
        if s.strip().startswith("|"):
            cells = [strip_md(c) for c in s.strip().strip("|").split("|")]
            if all(set(c.strip()) <= set("-: ") for c in cells):
                continue
            put("표 머리" if not table_hdr_done else "표", cells)
            table_hdr_done = True
            continue
        m = re.match(r"^-\s*\[( |x)\]\s*(.*)$", s.strip())
        if m:
            put("기록 규칙", [("[완료] " if m.group(1) == "x" else "[ ] ") +
                           strip_md(m.group(2))])
            continue
        if s.strip().startswith("- "):
            txt = strip_md(s.strip()[2:])
            warn = txt.startswith("⚠") or "함정" in txt or "오류" in txt or "실패" in txt
            put("주의" if warn else "설명", [txt], warn=warn)
            continue
        if s.strip().startswith(">"):
            put("메모", [strip_md(s.strip().lstrip("> "))])
            continue
        put("설명", [strip_md(s)])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:E{max(r - 1, 2)}"
    return r - 1


# ---------------------------------------------------------------------------
# ④ 확인 완료 이력
# ---------------------------------------------------------------------------


def build_done(wb, sec3, sec6, sec7, verify_meta, orphan_notes):
    ws = wb.create_sheet(SHEET_DONE)
    write_header(ws, 1, ["구분", "분류", "확정 사항 — 재검증 대상 아님", "출처", "비고"],
                 [16, 26, 96, 40, 28])
    r = 2
    ws.cell(r, 1, "여기 적힌 것은 원문 대조가 끝난 사실이다. 원고가 이와 다르게 쓰고 있다면 "
                  "원고가 틀린 것이다. 재검토 시 중복 작업을 막는 용도.").font = NOTE_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    r += 1

    def put(gubun, bunryu, text, src="", note=""):
        nonlocal r
        for c, v in enumerate([gubun, bunryu, text, src, note], 1):
            cell = ws.cell(r, c, v)
            cell.border = BOX
            cell.font = Font(size=10)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(r, 1).fill = PatternFill("solid", fgColor="EAF4E6")
        r += 1

    r = band(ws, r, "■ 근거 파일 — 05_review/verify_*.md (1차 출처 취득 성공분)", 5)
    for name, meta in verify_meta.items():
        put("근거 파일", name, meta.get("title", ""),
            meta.get("source", ""), meta.get("basis", ""))
    r += 1

    r = band(ws, r, "■ verify 확정 사항 — 원고 반영 확인 대상 (체크리스트 §3)", 5)
    cur = ""
    for ln in sec3:
        h = re.match(r"^###\s+(3\.\d+)\s*(.*)$", ln)
        if h:
            cur = f"{h.group(1)} {strip_md(h.group(2))}"
            continue
        m = re.match(r"^-\s*\[( |x)\]\s*(🔴|🟡|🟢)?\s*(.*)$", ln.strip())
        if m:
            put("§3 확정", cur, strip_md(m.group(3)),
                "law.go.kr 원문 (verify_*.md)", m.group(2) or "")
    r += 1

    r = band(ws, r, "■ 확인 완료 — 그대로 사용 가능한 사실 (체크리스트 §7)", 5)
    cur = ""
    for ln in sec7:
        h = re.match(r"^###\s+(7\.\d+)\s*(.*)$", ln)
        if h:
            cur = f"{h.group(1)} {strip_md(h.group(2))}"
            continue
        if ln.strip().startswith("- "):
            txt = strip_md(ln.strip()[2:])
            url = ""
            um = re.search(r"(https?://\S+|lx\.scourt\.go\.kr/\S+|law\.go\.kr/\S+)", txt)
            if um:
                url = um.group(1).rstrip(".")
            put("§7 확정", cur, txt, url or "law.go.kr 원문 (verify_*.md)", "")
    r += 1

    r = band(ws, r, "■ 기존 체크리스트에서 해소된 항목 (체크리스트 §6 — 이력)", 5)
    for ln in sec6:
        if not ln.strip().startswith("|"):
            continue
        cells = [strip_md(c) for c in ln.strip().strip("|").split("|")]
        if len(cells) < 4 or cells[0].strip() in ("#", "") or \
                all(set(c.strip()) <= set("-: ") for c in cells):
            continue
        put(f"§6 이력 #{cells[0].strip()}", cells[2].strip(), cells[1].strip(),
            cells[3].strip(), "")
    r += 1

    if orphan_notes:
        r = band(ws, r, "■ 구본(수동 제작본) 시트에 남아 있던 사람 기입분 — 보존", 5,
                 fill=PatternFill("solid", fgColor="FBD5B5"),
                 font=Font(bold=True, size=10, color="833C00"))
        for sheet, row, col, val in orphan_notes:
            put("구본 기입", f"{sheet} R{row}", val, col, "재생성 시 승계 키가 없어 이력으로 이관")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:E{max(r - 1, 2)}"
    return r - 1


def read_verify_meta(review_dir: str) -> "OrderedDict[str, dict]":
    out: "OrderedDict[str, dict]" = OrderedDict()
    for path in sorted(glob.glob(os.path.join(review_dir, "verify_*.md"))):
        name = os.path.basename(path)
        meta = {"title": "", "source": "", "basis": ""}
        with open(path, encoding="utf-8") as fh:
            for ln in fh:
                s = ln.strip()
                if s.startswith("# ") and not meta["title"]:
                    meta["title"] = strip_md(s[2:])
                elif s.startswith("검증 기준"):
                    meta["basis"] = strip_md(s.split(":", 1)[-1])
                elif s.startswith(("1차 출처", "검증 방법", "성공한 원문 경로")):
                    meta["source"] = strip_md(s.split(":", 1)[-1])
        if not meta["source"]:
            meta["source"] = f"05_review/{name} (파일 본문에 경로 기재)"
        out[name] = meta
    return out


# ---------------------------------------------------------------------------
# ⑤⑥ 레거시 승계
# ---------------------------------------------------------------------------


def build_legacy(wb, legacy: dict):
    made = {}
    for dst in SHEET_LEGACY:
        data = legacy.get(dst)
        if not data:
            continue
        ws = wb.create_sheet(dst)
        rows = data["rows"]
        for c, w in enumerate(data["widths"], 1):
            if w:
                ws.column_dimensions[get_column_letter(c)].width = w
        for ri, row in enumerate(rows, 1):
            for ci, v in enumerate(row, 1):
                cell = ws.cell(ri, ci, v)
                cell.border = BOX
                if ri == 1:
                    cell.fill, cell.font, cell.alignment = (
                        HEADER_FILL, HEADER_FONT, HEADER_ALIGN)
                else:
                    cell.font = Font(size=10)
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.freeze_panes = "A2"
        if len(rows) > 1:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(rows[0]))}{len(rows)}"
        made[dst] = len(rows) - 1
    return made


# ---------------------------------------------------------------------------
# 재검증
# ---------------------------------------------------------------------------


def verify_output(path: str, expected_items: int) -> bool:
    log("")
    log("=" * 78)
    log("[검증] 생성된 xlsx 를 openpyxl 로 다시 열어 확인")
    log("=" * 78)
    ok = True
    wb = openpyxl.load_workbook(path)
    log(f"  시트 {len(wb.sheetnames)}개: {wb.sheetnames}")

    ws = wb[SHEET_MAIN]
    data_rows = ws.max_row - 1
    log(f"  {SHEET_MAIN}: {data_rows}행 x {ws.max_column}열  (기대 {expected_items}행)")
    if data_rows != expected_items:
        log(f"  [FAIL] 행 수 불일치")
        ok = False
    else:
        log(f"  [OK] 체크리스트 마커 전량 등재 — {data_rows}건")

    log(f"  틀 고정: {ws.freeze_panes}  |  자동 필터: {ws.auto_filter.ref}")
    if ws.freeze_panes != "E2" or not ws.auto_filter.ref:
        log("  [FAIL] 틀고정/자동필터 미적용")
        ok = False
    else:
        log("  [OK] 틀 고정 + 자동 필터 적용")

    dvs = ws.data_validations.dataValidation
    if dvs:
        dv = dvs[0]
        log(f"  [OK] 드롭다운(데이터 검증): {dv.type} {dv.formula1} 적용범위 {dv.sqref}")
        for opt in STATUS_OPTIONS:
            if opt not in str(dv.formula1):
                log(f"  [FAIL] 드롭다운 항목 누락: {opt}")
                ok = False
    else:
        log("  [FAIL] 드롭다운 없음")
        ok = False

    cf = sum(len(rules) for rules in ws.conditional_formatting._cf_rules.values())
    log(f"  조건부 서식 규칙: {cf}개 (확인완료=녹색 / 수정필요=주황)")
    if cf < 2:
        log("  [FAIL] 조건부 서식 누락")
        ok = False

    widths = sum(1 for c in range(1, ws.max_column + 1)
                 if ws.column_dimensions[get_column_letter(c)].width)
    wraps = sum(1 for c in range(1, 15) if ws.cell(2, c).alignment.wrap_text)
    log(f"  열 너비 지정 {widths}열 / 줄바꿈 적용 {wraps}열")

    dist = defaultdict(int)
    blanks = defaultdict(int)
    status = defaultdict(int)
    no_marker = 0
    for r in range(2, ws.max_row + 1):
        dist[ws.cell(r, 4).value] += 1
        status[ws.cell(r, 9).value] += 1
        for c, name in ((3, "절·위치"), (5, "마커원문"), (7, "확인내용"), (15, "키"),
                        (16, "자동판정")):
            if ws.cell(r, c).value in (None, ""):
                blanks[name] += 1
        if not ws.cell(r, 5).value:
            no_marker += 1
    log("  우선순위 분포: " + " / ".join(f"{k} {v}" for k, v in
                                   sorted(dist.items(), key=lambda kv:
                                          PRIORITY_ORDER.index(kv[0]))))
    log("  상태 초기값 분포: " + " / ".join(f"{k} {v}" for k, v in status.items()))
    log(f"  빈 셀 점검: " + (", ".join(f"{k} {v}건" for k, v in blanks.items())
                          if blanks else "핵심 열 공백 0건"))
    if no_marker:
        log(f"  [주의] 마커 원문이 비어 있는 행 {no_marker}건")

    prio_fills = sum(1 for r in range(2, ws.max_row + 1)
                     if ws.cell(r, 1).fill and
                     ws.cell(r, 1).fill.fgColor.rgb not in (None, "00000000"))
    log(f"  우선순위별 행 색상 적용: {prio_fills}행")

    ws2 = wb[SHEET_SUMMARY]
    formulas = [ws2.cell(r, c).value for r in range(1, ws2.max_row + 1)
                for c in range(1, 11)
                if isinstance(ws2.cell(r, c).value, str)
                and ws2.cell(r, c).value.startswith("=")]
    log(f"  {SHEET_SUMMARY}: {ws2.max_row}행 / 수식 {len(formulas)}개")
    if len(formulas) < 50:
        log("  [FAIL] 요약 수식이 너무 적음")
        ok = False
    else:
        log(f"  [OK] 요약 수식 예시: {formulas[0]}")
        log(f"                    {formulas[6]}")

    bad = [f for f in formulas if SHEET_MAIN not in f and "SUM(" not in f
           and "IFERROR(" not in f]
    if bad:
        log(f"  [FAIL] 작업표를 참조하지 않는 수식 {len(bad)}개")
        ok = False
    else:
        log("  [OK] 모든 COUNTIF/COUNTIFS 수식이 ① 작업표를 참조 (상태 채우면 자동 갱신)")

    for name in (SHEET_HOWTO, SHEET_DONE):
        if name in wb.sheetnames:
            w = wb[name]
            log(f"  {name}: {w.max_row}행 x {w.max_column}열 (틀고정 {w.freeze_panes})")
        else:
            log(f"  [FAIL] 시트 없음: {name}")
            ok = False
    for name in SHEET_LEGACY:
        if name in wb.sheetnames:
            w = wb[name]
            log(f"  {name}: {w.max_row}행 (구본 승계)")
    wb.close()
    log("")
    log("[검증 결과] " + ("PASS — 모든 항목 통과" if ok else "FAIL — 위 [FAIL] 항목 확인"))
    return ok


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------


def main(argv: list[str]) -> int:
    args = [a for a in argv[1:] if not a.startswith("--")]
    dry = "--dry-run" in argv
    if not args:
        sys.stderr.write(__doc__)
        return 2

    ws_dir = os.path.abspath(args[0])
    if not os.path.isdir(ws_dir):
        sys.stderr.write(f"작업 폴더가 없습니다: {ws_dir}\n")
        return 2

    checklist = os.path.join(ws_dir, "02_verification_checklist.md")
    chapters = os.path.join(ws_dir, "chapters")
    review = os.path.join(ws_dir, "05_review")
    out = os.path.join(ws_dir, "원문대조_작업표.xlsx")
    if not os.path.exists(checklist):
        sys.stderr.write(f"체크리스트가 없습니다: {checklist}\n")
        return 2

    log("=" * 78)
    log(f"원문대조 작업표 생성 — {os.path.basename(ws_dir)}")
    log("=" * 78)
    log(f"  체크리스트: {checklist}")

    with open(checklist, encoding="utf-8") as fh:
        lines = fh.read().split("\n")

    chap_table = parse_chapter_table(
        slice_section(lines, r"^### 0\.3 ", r"^### 0\.4 "))
    focus = parse_bullets(slice_section(lines, r"^### 0\.4 ", r"^### 0\.5 "))
    common = parse_bullets(slice_section(lines, r"^### 0\.5 ", r"^## 1\. "))
    sec1 = slice_section(lines, r"^## 1\. ", r"^## 2\. ")
    watch = parse_bullets(slice_section(lines, r"^## 2\. ", r"^## 3\. "))
    sec3 = slice_section(lines, r"^## 3\. ", r"^## 4\. ")
    sec4 = slice_section(lines, r"^## 4\. ", r"^## 5\. ")
    sec5 = slice_section(lines, r"^## 5\. ", r"^## 6\. ")
    sec6 = slice_section(lines, r"^## 6\. ", r"^## 7\. ")
    sec7 = slice_section(lines, r"^## 7\. ", r"^## 8\. ")

    items4, extras = parse_section4(sec4)
    items5 = parse_section5(sec5)
    log(f"  §0.3 장별 분포 표: {len(chap_table)}개 장")
    log(f"  §4 장별 검증 항목: {len(items4)}건 / 보완 과제 {len(extras)}건")
    log(f"  §5 표기·상호참조 마커: {len(items5)}건")
    log(f"  §0.4 집중 검증 권고 {len(focus)}건 / §0.5 전 장 공통 {len(common)}건 / "
        f"§2 개정 감시 {len(watch)}건")

    items = items4 + items5
    for it in items:
        it.setdefault("표기용", False)
    dupe = defaultdict(int)
    for it in items:
        k = (it["장"], it["행"])
        dupe[k] += 1
        it["중복"] = dupe[k]

    markers = collect_manuscript_markers(chapters)
    total_markers = sum(len(v) for v in markers.values())
    log(f"  원고 마커([검증필요…]): {total_markers}건 "
        f"({len(glob.glob(os.path.join(chapters, '*.md')))}개 장 파일)")

    paired, exact = pair_with_manuscript(items, markers)
    log(f"  체크리스트 ↔ 원고 마커 짝짓기: {paired}/{len(items)}건 "
        f"(행번호 정확 일치 {exact}건)")

    items.sort(key=lambda r: (chapter_sort_key(r["장"]), r["행"] or 0, r["중복"]))
    for ch, meta in chap_table.items():
        got = sum(1 for it in items if it["장"] == ch)
        if got != meta["마커"]:
            log(f"  [주의] 제{ch}장 — 표 {meta['마커']}건 vs 파싱 {got}건")

    prior = defaultdict(int)
    for it in items:
        prior[it["우선순위"]] += 1
    log("  우선순위 분포: " + " / ".join(f"{p} {prior[p]}" for p in PRIORITY_ORDER))

    log("")
    log("[기존 파일] 사람 기입분 확인")
    existing = read_existing(out)

    verify_meta = read_verify_meta(review)
    log(f"  05_review/verify_*.md: {len(verify_meta)}개")

    if dry:
        log("")
        log("--dry-run — 파일을 쓰지 않고 종료")
        return 0

    log("")
    log("[생성] 시트 작성")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    restored = build_main(wb, items, existing["entries"])
    build_summary(wb, items, chap_table, focus, common, extras, watch, prior)
    n3 = build_howto(wb, sec1)
    n4 = build_done(wb, sec3, sec6, sec7, verify_meta, existing["orphan_notes"])
    made = build_legacy(wb, existing["legacy"])
    log(f"  {SHEET_MAIN}: {len(items)}행 (승계 {restored}행)")
    log(f"  {SHEET_HOWTO}: {n3}행 / {SHEET_DONE}: {n4}행")
    for k, v in made.items():
        log(f"  {k}: {v}행 (구본 승계)")

    wb.active = 0
    wb.save(out)
    log(f"  저장: {out}  ({os.path.getsize(out):,} bytes)")

    return 0 if verify_output(out, len(items)) else 1


if __name__ == "__main__":
    sys.exit(main(sys.argv))
