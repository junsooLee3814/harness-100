# -*- coding: utf-8 -*-
"""
build_xlsx.py — 사업수지 엑셀 워크북 빌더 (수식 내장)

사용:
    python build_xlsx.py <config.json> <output.xlsx>

설계 근거: _design/excel_schema.md (실제 PF 금융모델 분해 결과)
- 값이 아니라 '수식'을 기입한다. 계산은 Excel이 한다.
- 사용자가 00_입력 시트의 셀을 바꾸면 전 시트가 재계산된다.
- 모든 배분 행에 Total 열 + 균형검증 열(TRUE/FALSE)을 붙인다.
"""
import sys
import json

sys.stdout.reconfigure(encoding="utf-8", errors="replace")
sys.stderr.reconfigure(encoding="utf-8", errors="replace")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("[ERROR] openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# ---------------------------------------------------------------- 스타일
F_TITLE = Font(bold=True, size=14, color="1F3864")
F_HEAD = Font(bold=True, size=10, color="FFFFFF")
F_SUB = Font(bold=True, size=10)
F_NORM = Font(size=10)
FILL_HEAD = PatternFill("solid", fgColor="1F3864")
FILL_SUB = PatternFill("solid", fgColor="D9E2F3")
FILL_IN = PatternFill("solid", fgColor="FFF2CC")   # 사용자 입력 셀
FILL_OK = PatternFill("solid", fgColor="E2EFDA")
THIN = Side(style="thin", color="B4C6E7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
NUM = "#,##0"
PCT = "0.00%"


def head(ws, row, labels, start_col=1):
    for i, lab in enumerate(labels):
        c = ws.cell(row=row, column=start_col + i, value=lab)
        c.font = F_HEAD
        c.fill = FILL_HEAD
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
    return row + 1


def section(ws, row, title, span=8):
    c = ws.cell(row=row, column=1, value=title)
    c.font = F_SUB
    c.fill = FILL_SUB
    for i in range(1, span + 1):
        ws.cell(row=row, column=i).fill = FILL_SUB
        ws.cell(row=row, column=i).border = BORDER
    return row + 1


WARNINGS = []


def warn(msg):
    """경고를 기록하고 출력한다. 경고가 1건이라도 있으면 종료 코드 3으로 실패한다
    (자동화가 종료 코드만 보고 통과시키는 것을 막기 위함)."""
    WARNINGS.append(msg)
    print(f"[WARN] {msg}", file=sys.stderr)


def note(v):
    """설명·근거 텍스트가 '='로 시작하면 Excel이 수식으로 오인해 #NAME? 을 낸다.
    앞에 공백을 넣어 텍스트로 강제한다."""
    if isinstance(v, str) and v.startswith("="):
        return " " + v
    return v


def widths(ws, spec):
    for col, w in spec.items():
        ws.column_dimensions[col].width = w


# ---------------------------------------------------------------- 00_입력
def sheet_input(wb, cfg):
    ws = wb.create_sheet("00_입력")
    widths(ws, {"A": 22, "B": 26, "C": 16, "D": 12, "E": 40})
    ws["A1"] = f"{cfg['project']['name']} — 입력 변수"
    ws["A1"].font = F_TITLE
    r = 3

    r = section(ws, r, "■ 사업 개요")
    r = head(ws, r, ["구분", "항목", "값", "단위", "비고"])
    site = cfg["site"]
    rows = [
        ("사업", "사업유형", cfg["project"].get("type", ""), "", "정비사업 / 민간 자체시행"),
        ("", "소재지", cfg["project"].get("location", ""), "", "조례·분양가상한제 확인 대상"),
        ("규모", "대지면적", site.get("대지면적_m2", 0), "㎡", ""),
        ("", "대지면적(평)", None, "평", "㎡ × 0.3025 자동환산"),
        ("", "지상연면적", site.get("지상연면적_m2", 0), "㎡", ""),
        ("", "지상연면적(평)", None, "평", "㎡ × 0.3025 자동환산"),
        ("", "지하연면적", site.get("지하연면적_m2", 0), "㎡", ""),
        ("", "연면적 합계(평)", None, "평", "(지상+지하) × 0.3025 자동환산"),
        ("", "건폐율", site.get("건폐율", 0), "", ""),
        ("", "용적률", site.get("용적률", 0), "", ""),
        ("", "세대수", site.get("세대수", 0), "세대", ""),
        ("", "주차대수", site.get("주차대수", 0), "대", ""),
    ]
    ref = {}
    for grp, item, val, unit, note_txt in rows:
        ws.cell(row=r, column=1, value=grp).font = F_NORM
        ws.cell(row=r, column=2, value=item).font = F_NORM
        c = ws.cell(row=r, column=3)
        if item == "대지면적(평)":
            c.value = f"=C{r-1}*0.3025"
        elif item == "지상연면적(평)":
            c.value = f"=C{r-1}*0.3025"
        elif item == "연면적 합계(평)":
            c.value = f"=(C{r-3}+C{r-1})*0.3025"
        else:
            c.value = val
            c.fill = FILL_IN
        c.number_format = NUM if isinstance(val, (int, float)) or val is None else "General"
        ws.cell(row=r, column=4, value=unit)
        ws.cell(row=r, column=5, value=note_txt).font = F_NORM
        for i in range(1, 6):
            ws.cell(row=r, column=i).border = BORDER
        ref[item] = f"'00_입력'!$C${r}"
        r += 1

    r += 1
    r = section(ws, r, "■ 사업 기간")
    r = head(ws, r, ["구분", "항목", "값", "단위", "비고"])
    per = cfg["period"]
    for grp, label, key, note_txt in [
        ("기간", "인허가 기간", "인허가_개월", "표준 일정"),
        ("", "공사 기간", "공사_개월", ""),
        ("", "분양~정산 기간", "분양정산_개월", ""),
        ("", "지연 버퍼", "지연버퍼_개월", "★ 별도 계상 — 은폐 금지"),
    ]:
        ws.cell(row=r, column=1, value=grp)
        ws.cell(row=r, column=2, value=label)
        c = ws.cell(row=r, column=3, value=per.get(key, 0))
        c.fill = FILL_IN
        ws.cell(row=r, column=4, value="개월")
        ws.cell(row=r, column=5, value=note_txt).font = F_NORM
        for i in range(1, 6):
            ws.cell(row=r, column=i).border = BORDER
        ref[key] = f"'00_입력'!$C${r}"
        r += 1
    ws.cell(row=r, column=2, value="총 사업기간").font = F_SUB
    ws.cell(row=r, column=3, value=f"=SUM(C{r-4}:C{r-1})").font = F_SUB
    ws.cell(row=r, column=4, value="개월")
    ref["총사업기간"] = f"'00_입력'!$C${r}"
    for i in range(1, 6):
        ws.cell(row=r, column=i).border = BORDER
    r += 2

    r = section(ws, r, "■ 금융 조건")
    r = head(ws, r, ["구분", "항목", "값", "단위", "비고"])
    fund = cfg.get("funding", {})
    ws.cell(row=r, column=2, value="자기자본(Equity)")
    c = ws.cell(row=r, column=3, value=fund.get("equity_천원", 0))
    c.fill = FILL_IN
    c.number_format = NUM
    ws.cell(row=r, column=4, value="천원")
    ref["equity"] = f"'00_입력'!$C${r}"
    for i in range(1, 6):
        ws.cell(row=r, column=i).border = BORDER
    r += 1
    for t in fund.get("tranches", []):
        for label, key, unit, fmt in [
            (f"{t['구분']} 한도", "한도_천원", "천원", NUM),
            (f"{t['구분']} 금리", "금리", "", PCT),
            (f"{t['구분']} 취급수수료", "취급수수료", "", PCT),
        ]:
            ws.cell(row=r, column=2, value=label)
            c = ws.cell(row=r, column=3, value=t.get(key, 0))
            c.fill = FILL_IN
            c.number_format = fmt
            ws.cell(row=r, column=4, value=unit)
            for i in range(1, 6):
                ws.cell(row=r, column=i).border = BORDER
            ref[f"{t['구분']}_{key}"] = f"'00_입력'!$C${r}"
            r += 1
    r += 1

    r = section(ws, r, "■ 분양 대금 납입 구조")
    r = head(ws, r, ["구분", "항목", "값", "단위", "비고"])
    pay = cfg.get("schedule", {}).get("납입", {})
    for label, key, note_txt in [
        ("계약금 비율", "계약금", ""),
        ("중도금 차수", "중도금차수", "회"),
        ("중도금 회차별 비율", "중도금비율", ""),
        ("잔금 비율", "잔금", ""),
        ("적립계좌 비율(준공전)", "적립비율_준공전", "대출 상환 재원"),
        ("적립계좌 비율(준공후)", "적립비율_준공후", ""),
    ]:
        ws.cell(row=r, column=2, value=label)
        c = ws.cell(row=r, column=3, value=pay.get(key, 0))
        c.fill = FILL_IN
        c.number_format = PCT if key not in ("중도금차수",) else "0"
        ws.cell(row=r, column=5, value=note_txt).font = F_NORM
        for i in range(1, 6):
            ws.cell(row=r, column=i).border = BORDER
        ref[f"납입_{key}"] = f"'00_입력'!$C${r}"
        r += 1

    ws.cell(row=r + 1, column=1,
            value="※ 노란색 셀만 수정하십시오. 나머지는 수식으로 연결되어 자동 재계산됩니다.").font = F_NORM
    return ref


# ---------------------------------------------------------------- 01_사업수지
def sheet_pl(wb, cfg, ref):
    ws = wb.create_sheet("01_사업수지")
    widths(ws, {"A": 12, "B": 26, "C": 18, "D": 14, "E": 16, "F": 10, "G": 20,
                "H": 16, "I": 16, "J": 16, "K": 14, "L": 14})
    ws["A1"] = f"{cfg['project']['name']} — 사업수지"
    ws["A1"].font = F_TITLE
    ws["A2"] = "(단위: 천원, VAT 별도)"
    r = 4
    idx = {}

    # ---- 수입
    r = section(ws, r, "■ 수입", 7)
    r = head(ws, r, ["구분", "세부", "분양면적(평)", "평당분양가", "수입액", "구성비", "비고"])
    rev_start = r
    for it in cfg.get("revenue", []):
        ws.cell(row=r, column=1, value=note(it.get("구분", "")))
        ws.cell(row=r, column=2, value=note(it.get("세부", "")))
        ws.cell(row=r, column=3, value=it.get("분양면적_평", 0)).number_format = NUM
        ws.cell(row=r, column=4, value=it.get("평당분양가_천원", 0)).number_format = NUM
        ws.cell(row=r, column=5, value=f"=C{r}*D{r}").number_format = NUM
        ws.cell(row=r, column=6, value=f"=IFERROR(E{r}/$E${{TOT}},0)").number_format = PCT
        ws.cell(row=r, column=7, value=note(it.get("근거", "")))
        for i in range(1, 8):
            ws.cell(row=r, column=i).border = BORDER
        r += 1
    rev_end = r - 1
    ws.cell(row=r, column=2, value="분양 총액").font = F_SUB
    ws.cell(row=r, column=3, value=f"=SUM(C{rev_start}:C{rev_end})").number_format = NUM
    ws.cell(row=r, column=5, value=f"=SUM(E{rev_start}:E{rev_end})").number_format = NUM
    idx["분양총액"] = r
    r += 1
    ws.cell(row=r, column=2, value="VAT")
    ws.cell(row=r, column=5, value=cfg.get("vat_천원", 0)).number_format = NUM
    idx["vat"] = r
    r += 1
    ws.cell(row=r, column=2, value="수입 총계").font = F_SUB
    ws.cell(row=r, column=5, value=f"=E{idx['분양총액']}+E{idx['vat']}").number_format = NUM
    ws.cell(row=r, column=5).font = F_SUB
    idx["수입총계"] = r
    for i in range(1, 8):
        ws.cell(row=r, column=i).border = BORDER
    # 구성비 분모 치환
    for rr in range(rev_start, rev_end + 1):
        ws.cell(row=rr, column=6).value = f"=IFERROR(E{rr}/$E${idx['수입총계']},0)"
    r += 2

    # ---- 지출
    r = section(ws, r, "■ 지출", 12)
    r = head(ws, r, ["분류", "항목", "산출근거", "수량", "단가·요율", "구성비", "신뢰도",
                     "금액", "기투입(Equity)", "필수(PF)", "분양불", "준공후불"])
    cost_hdr = r - 1
    group_rows = {}
    for grp in ["토지비", "건축비", "부대비용", "금융비용"]:
        items = cfg.get("cost", {}).get(grp, [])
        if not items:
            continue
        start = r
        for it in items:
            ws.cell(row=r, column=1, value=grp if r == start else "")
            ws.cell(row=r, column=2, value=note(it.get("항목", "")))
            ws.cell(row=r, column=3, value=note(it.get("산출근거", "")))
            qty = it.get("수량")
            rate = it.get("단가")
            amt = it.get("금액_천원")
            if qty is not None and rate is not None:
                ws.cell(row=r, column=4, value=qty).number_format = NUM
                ws.cell(row=r, column=5, value=rate).number_format = NUM if rate >= 1 else PCT
                ws.cell(row=r, column=8, value=f"=D{r}*E{r}").number_format = NUM
            else:
                ws.cell(row=r, column=8, value=amt or 0).number_format = NUM
            ws.cell(row=r, column=7, value=note(it.get("신뢰도", "")))
            # 배분율: 필수 + 분양불 + 준공후불 = 1 이어야 한다.
            # 필수비율 미지정 시 나머지로 자동 산정 (합계 200% 오류 방지)
            x = float(it.get("분양불비율", 0.0))
            y = float(it.get("준공후불비율", 0.0))
            w = it.get("필수비율")
            w = (1.0 - x - y) if w is None else float(w)
            total_ratio = round(w + x + y, 6)
            if abs(total_ratio - 1.0) > 1e-6:
                warn(f"배분율 합계 {total_ratio} != 1 — {grp}/{it.get('항목')} "
                     f"(필수 {w}, 분양불 {x}, 준공후불 {y})")
            eq = it.get("기투입_천원", 0)
            ws.cell(row=r, column=9, value=eq).number_format = NUM
            ws.cell(row=r, column=10, value=f"=($H{r}-$I{r})*{w}").number_format = NUM
            ws.cell(row=r, column=11, value=f"=($H{r}-$I{r})*{x}").number_format = NUM
            ws.cell(row=r, column=12, value=f"=($H{r}-$I{r})*{y}").number_format = NUM
            if it.get("비례율제외"):
                idx.setdefault("비례율제외행", []).append(r)
                idx.setdefault("비례율제외명", []).append(it.get("항목", ""))
                ws.cell(row=r, column=7).value = str(ws.cell(row=r, column=7).value or "") + " [비례율 제외]"
            for i in range(1, 13):
                ws.cell(row=r, column=i).border = BORDER
            r += 1
        ws.cell(row=r, column=2, value=f"{grp} 소계").font = F_SUB
        for col in (8, 9, 10, 11, 12):
            L = get_column_letter(col)
            ws.cell(row=r, column=col, value=f"=SUM({L}{start}:{L}{r-1})").number_format = NUM
            ws.cell(row=r, column=col).font = F_SUB
        for i in range(1, 13):
            ws.cell(row=r, column=i).border = BORDER
        group_rows[grp] = r
        r += 1

    ws.cell(row=r, column=2, value="지출 총계").font = F_SUB
    # 비용 그룹이 하나도 없으면 빈 수식('=')이 기입되어 파일이 손상된다 → 0으로 대체
    tot = "+".join(f"H{v}" for v in group_rows.values()) or "0"
    ws.cell(row=r, column=8, value=f"={tot}").number_format = NUM
    ws.cell(row=r, column=8).font = F_SUB
    for col, letter in [(9, "I"), (10, "J"), (11, "K"), (12, "L")]:
        s = "+".join(f"{letter}{v}" for v in group_rows.values()) or "0"
        ws.cell(row=r, column=col, value=f"={s}").number_format = NUM
    idx["지출총계"] = r
    for i in range(1, 13):
        ws.cell(row=r, column=i).border = BORDER
    # 구성비
    for grp, gr in group_rows.items():
        ws.cell(row=gr, column=6, value=f"=IFERROR(H{gr}/$H${idx['지출총계']},0)").number_format = PCT
    r += 1

    ws.cell(row=r, column=2, value="자금집행 균형 검증").font = F_SUB
    ws.cell(row=r, column=8,
            value=f"=IF(ROUND(H{idx['지출총계']},0)=ROUND(I{idx['지출총계']}+J{idx['지출총계']}"
                  f"+K{idx['지출총계']}+L{idx['지출총계']},0),\"OK\",\"불일치\")")
    ws.cell(row=r, column=8).fill = FILL_OK
    idx["집행검증"] = r
    r += 2

    # ---- 사업이익
    r = section(ws, r, "■ 사업이익", 7)
    for label, formula, fmt in [
        ("수입 총계", f"=E{idx['수입총계']}", NUM),
        ("지출 총계", f"=H{idx['지출총계']}", NUM),
        ("사업이익", f"=E{idx['수입총계']}-H{idx['지출총계']}", NUM),
        ("사업이익률 (이익/수입)", None, PCT),
        ("Profit on Cost (이익/총사업비)", None, PCT),
    ]:
        ws.cell(row=r, column=2, value=label).font = F_SUB
        if label == "사업이익률 (이익/수입)":
            ws.cell(row=r, column=5, value=f"=IFERROR(E{r-1}/E{r-3},0)")
        elif label.startswith("Profit"):
            ws.cell(row=r, column=5, value=f"=IFERROR(E{r-2}/E{r-3},0)")
        else:
            ws.cell(row=r, column=5, value=formula)
        ws.cell(row=r, column=5).number_format = fmt
        ws.cell(row=r, column=5).font = F_SUB
        for i in range(1, 8):
            ws.cell(row=r, column=i).border = BORDER
        if label == "사업이익":
            idx["사업이익"] = r
        if label == "사업이익률 (이익/수입)":
            idx["이익률"] = r
        r += 1
    return idx


# ---------------------------------------------------------------- 02_현금흐름
def sheet_cf(wb, cfg, ref, pl):
    ws = wb.create_sheet("02_현금흐름")
    months = int(cfg.get("schedule", {}).get("총개월", 0)) or (
        int(cfg["period"].get("인허가_개월", 0)) + int(cfg["period"].get("공사_개월", 0))
        + int(cfg["period"].get("분양정산_개월", 0)) + int(cfg["period"].get("지연버퍼_개월", 0)))
    MAX_MONTHS = 180
    if months > MAX_MONTHS:
        warn(f"총 사업기간 {months}개월이 상한 {MAX_MONTHS}개월을 초과하여 절단했습니다. "
             f"기간을 나누어 검토하십시오.")
        months = MAX_MONTHS
    if months < 6:
        warn(f"총 사업기간이 {months}개월로 산정되어 최소 6개월로 보정했습니다. "
             f"period 블록을 확인하십시오.")
        months = 6
    prog = cfg.get("schedule", {}).get("공정률", [])
    sale = cfg.get("schedule", {}).get("분양률", [])

    # 드라이버 합계 검증 — 공정률은 100%여야 하고, 분양률은 100%를 넘을 수 없다.
    ps, ss = round(sum(prog), 6), round(sum(sale), 6)
    if abs(ps - 1.0) > 1e-6:
        warn(f"공정률 합계가 {ps:.1%} 입니다 (100%여야 함). "
             f"건축비 월별 배분이 총액과 불일치합니다.")
    if ss > 1.0 + 1e-6:
        warn(f"분양률 합계가 {ss:.1%} 로 100%를 초과합니다.")
    elif ss < 1.0 - 1e-6:
        warn(f"분양률 합계가 {ss:.1%} 입니다. 사업기간 내 미분양 "
             f"{1-ss:.1%} 가 남으며, 그만큼 분양수입이 회수되지 않습니다.")
    C0 = 4  # 첫 월 컬럼 = D

    ws["A1"] = f"{cfg['project']['name']} — 월별 현금흐름"
    ws["A1"].font = F_TITLE
    widths(ws, {"A": 20, "B": 22, "C": 16})
    for m in range(months):
        ws.column_dimensions[get_column_letter(C0 + m)].width = 13
    tcol = get_column_letter(C0 + months)      # Total
    vcol = get_column_letter(C0 + months + 1)  # 검증
    ws.column_dimensions[tcol].width = 15
    ws.column_dimensions[vcol].width = 10

    r = 3
    ws.cell(row=r, column=1, value="구분").font = F_HEAD
    ws.cell(row=r, column=2, value="항목").font = F_HEAD
    ws.cell(row=r, column=3, value="총액").font = F_HEAD
    for i in range(1, 4):
        ws.cell(row=r, column=i).fill = FILL_HEAD
    for m in range(months):
        c = ws.cell(row=r, column=C0 + m, value=f"{m+1}개월차")
        c.font = F_HEAD
        c.fill = FILL_HEAD
        c.alignment = Alignment(horizontal="center")
    ws.cell(row=r, column=C0 + months, value="Total").font = F_HEAD
    ws.cell(row=r, column=C0 + months).fill = FILL_HEAD
    ws.cell(row=r, column=C0 + months + 1, value="검증").font = F_HEAD
    ws.cell(row=r, column=C0 + months + 1).fill = FILL_HEAD
    r += 1

    def drive(row_label, values, fmt=PCT, cum=False):
        nonlocal r
        ws.cell(row=r, column=2, value=row_label).font = F_SUB
        for m in range(months):
            v = values[m] if m < len(values) else 0
            c = ws.cell(row=r, column=C0 + m, value=v)
            c.number_format = fmt
            c.fill = FILL_IN
        ws.cell(row=r, column=C0 + months,
                value=f"=SUM({get_column_letter(C0)}{r}:{get_column_letter(C0+months-1)}{r})").number_format = fmt
        this = r
        r += 1
        if cum:
            ws.cell(row=r, column=2, value=f"{row_label} 누계")
            for m in range(months):
                L = get_column_letter(C0 + m)
                if m == 0:
                    ws.cell(row=r, column=C0 + m, value=f"={L}{this}").number_format = fmt
                else:
                    P = get_column_letter(C0 + m - 1)
                    ws.cell(row=r, column=C0 + m, value=f"={P}{r}+{L}{this}").number_format = fmt
            cumr = r
            r += 1
            return this, cumr
        return this, None

    ws.cell(row=r, column=1, value="진행률 드라이버").font = F_SUB
    r += 1
    prog_r, prog_c = drive("공정률", prog, PCT, cum=True)
    sale_r, sale_c = drive("분양률", sale, PCT, cum=True)
    r += 1

    # ---- 수입
    ws.cell(row=r, column=1, value="수입").font = F_SUB
    r += 1
    pay = cfg.get("schedule", {}).get("납입", {})
    n_mid = int(pay.get("중도금차수", 4) or 0)
    tranche_pay = [("계약금", pay.get("계약금", 0.1), 0)]
    gap = max(1, months // (n_mid + 2)) if n_mid else 1
    for i in range(n_mid):
        tranche_pay.append((f"중도금 {i+1}차", pay.get("중도금비율", 0.1), (i + 1) * gap))
    tranche_pay.append(("잔금", pay.get("잔금", 0.3), months - 1))

    rev_rows = []
    for name, ratio, start_m in tranche_pay:
        ws.cell(row=r, column=2, value=name)
        ws.cell(row=r, column=3, value=f"='01_사업수지'!E{pl['분양총액']}*{ratio}").number_format = NUM
        for m in range(months):
            L = get_column_letter(C0 + m)
            if m < start_m:
                continue
            if m == start_m:
                # 도래 시점: 누계 분양률 기준 일괄
                ws.cell(row=r, column=C0 + m, value=f"=$C{r}*{L}{sale_c}").number_format = NUM
            else:
                ws.cell(row=r, column=C0 + m, value=f"=$C{r}*{L}{sale_r}").number_format = NUM
        ws.cell(row=r, column=C0 + months,
                value=f"=SUM({get_column_letter(C0)}{r}:{get_column_letter(C0+months-1)}{r})").number_format = NUM
        ws.cell(row=r, column=C0 + months + 1, value=f"=IF(ROUND({tcol}{r},0)<=ROUND(C{r},0)+1,\"OK\",\"초과\")")
        rev_rows.append(r)
        r += 1
    ws.cell(row=r, column=2, value="수입 계").font = F_SUB
    for m in range(months):
        L = get_column_letter(C0 + m)
        ws.cell(row=r, column=C0 + m,
                value=f"=SUM({L}{rev_rows[0]}:{L}{rev_rows[-1]})").number_format = NUM
    ws.cell(row=r, column=3, value=f"=SUM(C{rev_rows[0]}:C{rev_rows[-1]})").number_format = NUM
    ws.cell(row=r, column=C0 + months,
            value=f"=SUM({get_column_letter(C0)}{r}:{get_column_letter(C0+months-1)}{r})").number_format = NUM
    rev_tot = r
    r += 2

    # ---- 지출
    ws.cell(row=r, column=1, value="지출").font = F_SUB
    r += 1
    out_rows = []
    for label, src, mode in [
        ("토지비", "토지비", "front"),
        ("건축비", "건축비", "progress"),
        ("부대비용", "부대비용", "even"),
    ]:
        ws.cell(row=r, column=2, value=label)
        _sub = pl.get('소계_' + src)
        ws.cell(row=r, column=3,
                value=(f"='01_사업수지'!H{_sub}" if _sub else 0)).number_format = NUM
        for m in range(months):
            L = get_column_letter(C0 + m)
            if mode == "front" and m == 0:
                ws.cell(row=r, column=C0 + m, value=f"=$C{r}").number_format = NUM
            elif mode == "progress":
                ws.cell(row=r, column=C0 + m, value=f"=$C{r}*{L}{prog_r}").number_format = NUM
            elif mode == "even":
                ws.cell(row=r, column=C0 + m, value=f"=$C{r}/{months}").number_format = NUM
        ws.cell(row=r, column=C0 + months,
                value=f"=SUM({get_column_letter(C0)}{r}:{get_column_letter(C0+months-1)}{r})").number_format = NUM
        ws.cell(row=r, column=C0 + months + 1,
                value=f"=IF(ROUND({tcol}{r},0)=ROUND(C{r},0),\"OK\",\"불일치\")")
        out_rows.append(r)
        r += 1
    ws.cell(row=r, column=2, value="사업비 지출 계 (금융비 제외)").font = F_SUB
    for m in range(months):
        L = get_column_letter(C0 + m)
        ws.cell(row=r, column=C0 + m,
                value=f"=SUM({L}{out_rows[0]}:{L}{out_rows[-1]})").number_format = NUM
    out_tot = r
    r += 2

    # ---- 자금조달
    ws.cell(row=r, column=1, value="자금조달").font = F_SUB
    r += 1
    tr = (cfg.get("funding", {}).get("tranches") or [{"구분": "Tr.A", "금리": 0.08}])[0]
    rate_ref = ref.get(f"{tr['구분']}_금리", "0.08")

    ws.cell(row=r, column=2, value="기초 PF잔액")
    bal_open = r
    r += 1
    ws.cell(row=r, column=2, value="이자 (전월잔액×금리/12)")
    interest = r
    r += 1
    ws.cell(row=r, column=2, value="상환 (적립계좌 재원)")
    repay = r
    r += 1
    ws.cell(row=r, column=2, value="인출 (부족액 자동)")
    draw = r
    r += 1
    ws.cell(row=r, column=2, value="기말 PF잔액").font = F_SUB
    bal_close = r
    r += 1
    ws.cell(row=r, column=2, value="현금 잔액").font = F_SUB
    cash = r
    r += 1

    dep_pre = ref.get("납입_적립비율_준공전", "0.6")
    dep_post = ref.get("납입_적립비율_준공후", dep_pre)
    equity = ref.get("equity", "0")
    # 준공 시점 = 인허가 + 공사 기간. 이후에는 준공후 적립비율을 적용한다.
    done_m = int(cfg["period"].get("인허가_개월", 0)) + int(cfg["period"].get("공사_개월", 0))
    for m in range(months):
        L = get_column_letter(C0 + m)
        P = get_column_letter(C0 + m - 1) if m > 0 else None
        ws.cell(row=bal_open, column=C0 + m,
                value="=0" if m == 0 else f"={P}{bal_close}").number_format = NUM
        ws.cell(row=interest, column=C0 + m,
                value=f"={L}{bal_open}*{rate_ref}/12").number_format = NUM
        _dep = dep_post if m >= done_m else dep_pre
        ws.cell(row=repay, column=C0 + m,
                value=f"=MIN({L}{bal_open},{L}{rev_tot}*{_dep})").number_format = NUM
        prev_cash = f"{P}{cash}" if m > 0 else equity
        ws.cell(row=draw, column=C0 + m,
                value=f"=MAX(0,-({prev_cash}+{L}{rev_tot}-{L}{out_tot}-{L}{interest}-{L}{repay}))"
                ).number_format = NUM
        ws.cell(row=bal_close, column=C0 + m,
                value=f"={L}{bal_open}+{L}{draw}-{L}{repay}").number_format = NUM
        ws.cell(row=cash, column=C0 + m,
                value=f"={prev_cash}+{L}{rev_tot}-{L}{out_tot}-{L}{interest}+{L}{draw}-{L}{repay}"
                ).number_format = NUM
    ws.cell(row=interest, column=C0 + months,
            value=f"=SUM({get_column_letter(C0)}{interest}:{get_column_letter(C0+months-1)}{interest})"
            ).number_format = NUM
    ws.cell(row=draw, column=C0 + months,
            value=f"=MAX({get_column_letter(C0)}{draw}:{get_column_letter(C0+months-1)}{draw})"
            ).number_format = NUM
    ws.cell(row=draw, column=2).value = "인출 (부족액 자동) / Total열=피크 인출"
    r += 1
    ws.cell(row=r, column=2, value="★ 피크자금 소요 (PF잔액 최대)").font = F_SUB
    ws.cell(row=r, column=3,
            value=f"=MAX({get_column_letter(C0)}{bal_close}:{get_column_letter(C0+months-1)}{bal_close})"
            ).number_format = NUM
    ws.cell(row=r, column=3).font = F_SUB
    peak = r
    r += 1
    ws.cell(row=r, column=2, value="현금 최저점 (음수면 조달 부족)").font = F_SUB
    ws.cell(row=r, column=3,
            value=f"=MIN({get_column_letter(C0)}{cash}:{get_column_letter(C0+months-1)}{cash})"
            ).number_format = NUM
    lowcash = r
    r += 1
    ws.cell(row=r, column=2, value="PF 한도 초과 여부").font = F_SUB
    limits = "+".join(
        ref[f"{t['구분']}_한도_천원"] for t in cfg.get("funding", {}).get("tranches", [])
        if f"{t['구분']}_한도_천원" in ref) or "0"
    ws.cell(row=r, column=3, value=f"=IF(C{peak}>({limits}),\"한도 초과\",\"한도 내\")")
    ws.cell(row=r, column=3).fill = FILL_OK
    overlimit = r
    r += 1
    ws.cell(row=r, column=2, value="총 금융비용 (이자 합계)").font = F_SUB
    ws.cell(row=r, column=3, value=f"={tcol}{interest}").number_format = NUM
    fin_cost = r
    return {"이자합계": fin_cost, "피크자금": peak, "현금최저": lowcash,
            "한도초과": overlimit, "현금행": cash, "months": months,
            "수입계행": rev_tot, "총액열": tcol}


# ---------------------------------------------------------------- 03_민감도
def sheet_sens(wb, cfg, pl):
    ws = wb.create_sheet("03_민감도")
    widths(ws, {"A": 26, "B": 16, "C": 16, "D": 16, "E": 16, "F": 16})
    ws["A1"] = "민감도 분석"
    ws["A1"].font = F_TITLE
    r = 3
    ws.cell(row=r, column=1, value="기준 사업이익").font = F_SUB
    ws.cell(row=r, column=2, value=f"='01_사업수지'!E{pl['사업이익']}").number_format = NUM
    base = r
    r += 2

    r = section(ws, r, "■ 분양가 × 공사비 민감도 (사업이익, 천원)", 6)
    ws.cell(row=r, column=1, value="분양가 \\ 공사비")
    deltas_c = [-0.10, 0.0, 0.10, 0.20]
    deltas_p = [0.10, 0.0, -0.10]
    for j, dc in enumerate(deltas_c):
        ws.cell(row=r, column=2 + j, value=f"{dc:+.0%}").font = F_SUB
    r += 1
    rev_tot = f"'01_사업수지'!$E${pl['수입총계']}"
    cost_tot = f"'01_사업수지'!$H${pl['지출총계']}"
    _b = pl.get('소계_건축비')
    bld = f"'01_사업수지'!$H${_b}" if _b else "0"
    for dp in deltas_p:
        ws.cell(row=r, column=1, value=f"분양가 {dp:+.0%}").font = F_SUB
        for j, dc in enumerate(deltas_c):
            ws.cell(row=r, column=2 + j,
                    value=f"={rev_tot}*(1+{dp})-({cost_tot}+{bld}*{dc})").number_format = NUM
            ws.cell(row=r, column=2 + j).border = BORDER
        r += 1
    r += 1

    r = section(ws, r, "■ 손익분기", 6)
    ws.cell(row=r, column=1, value="손익분기 분양률")
    ws.cell(row=r, column=2, value=f"=IFERROR({cost_tot}/{rev_tot},0)").number_format = PCT
    ws.cell(row=r, column=3, value="총사업비 ÷ 총수입 — 이 비율만큼 팔려야 본전")
    be_row = r
    r += 1
    ws.cell(row=r, column=1, value="손익분기 평당 분양가")
    ws.cell(row=r, column=2,
            value=f"=IFERROR({cost_tot}/'01_사업수지'!C{pl['분양총액']},0)").number_format = NUM
    r += 2

    r = section(ws, r, "■ 분양률 민감도 (회수 기준 사업이익, 천원)", 6)
    ws.cell(row=r, column=1, value="분양률").font = F_SUB
    rates = [1.0, 0.9, 0.8, 0.7]
    for j2, s2 in enumerate(rates):
        ws.cell(row=r, column=2 + j2, value=f"{s2:.0%}").font = F_SUB
    r += 1
    ws.cell(row=r, column=1, value="회수 수입")
    for j2, s2 in enumerate(rates):
        ws.cell(row=r, column=2 + j2, value=f"={rev_tot}*{s2}").number_format = NUM
    r += 1
    ws.cell(row=r, column=1, value="사업이익")
    for j2, s2 in enumerate(rates):
        ws.cell(row=r, column=2 + j2, value=f"={rev_tot}*{s2}-{cost_tot}").number_format = NUM
        ws.cell(row=r, column=2 + j2).border = BORDER
    r += 2

    ws.cell(row=r, column=1,
            value="※ 공사비 단가가 [가정]이면 ±20% 구간을 반드시 검토하십시오.").font = F_NORM
    r += 1
    ws.cell(row=r, column=1,
            value="※ 금리 민감도는 별도 표를 두지 않습니다. 00_입력 시트의 금리 셀을 바꾸면 "
                  "02_현금흐름의 이자·PF잔액과 01_사업수지의 사업이익이 전부 재계산됩니다.").font = F_NORM
    r += 1
    ws.cell(row=r, column=1,
            value="※ 분양률 민감도는 총액 기준 근사입니다. 회수 시점 변화는 02_현금흐름의 "
                  "분양률 배열을 수정해 확인하십시오.").font = F_NORM
    return {"손익분기분양률": be_row}


# ---------------------------------------------------------------- 04_요약 / 05_검증
def sheet_summary(wb, cfg, pl, cf, ref, sens):
    ws = wb.create_sheet("04_요약")
    widths(ws, {"A": 30, "B": 22, "C": 46})
    ws["A1"] = f"{cfg['project']['name']} — 요약"
    ws["A1"].font = F_TITLE
    r = 3
    r = head(ws, r, ["항목", "값", "비고"])
    be = f"='03_민감도'!B{sens['손익분기분양률']}"
    limits = "+".join(
        ref[f"{t['구분']}_한도_천원"] for t in cfg.get("funding", {}).get("tranches", [])
        if f"{t['구분']}_한도_천원" in ref) or "0"
    rows = [
        ("사업유형", cfg["project"].get("type", ""), "", None),
        ("소재지", cfg["project"].get("location", ""), "", None),
        ("총 사업기간(개월)", f"={ref['총사업기간']}", "지연 버퍼 포함", "0"),
        ("수입 총계", f"='01_사업수지'!E{pl['수입총계']}", "", NUM),
        ("지출 총계", f"='01_사업수지'!H{pl['지출총계']}", "", NUM),
        ("사업이익", f"='01_사업수지'!E{pl['사업이익']}", "", NUM),
        ("사업이익률", f"='01_사업수지'!E{pl['이익률']}", "", PCT),
        ("총 금융비용", f"='02_현금흐름'!C{cf['이자합계']}", "월별 이자 합계", NUM),
        ("피크자금 소요", f"='02_현금흐름'!C{cf['피크자금']}", "PF 잔액 최대치", NUM),
        ("현금 최저점", f"='02_현금흐름'!C{cf['현금최저']}", "음수면 조달 부족", NUM),
        ("PF 한도 초과 여부", f"='02_현금흐름'!C{cf['한도초과']}", "", None),
        ("손익분기 분양률", be, "이 비율만큼 팔려야 본전", PCT),
        ("LTV (대출한도/수입총계)",
         (f"=({limits})/'01_사업수지'!E{pl['수입총계']}" if limits != "0" else 0), "", PCT),
        ("LTC (대출한도/총사업비)",
         (f"=({limits})/'01_사업수지'!H{pl['지출총계']}" if limits != "0" else 0), "", PCT),
    ]
    for a, b, c, fmt in rows:
        ws.cell(row=r, column=1, value=a).font = F_SUB
        cell = ws.cell(row=r, column=2, value=b)
        if fmt:
            cell.number_format = fmt
        ws.cell(row=r, column=3, value=c)
        for i in range(1, 4):
            ws.cell(row=r, column=i).border = BORDER
        r += 1

    ws2 = wb.create_sheet("05_검증")
    widths(ws2, {"A": 34, "B": 18, "C": 50})
    ws2["A1"] = "정합성 검증"
    ws2["A1"].font = F_TITLE
    r = 3
    r = head(ws2, r, ["검증 항목", "결과", "설명"])
    checks = [
        ("자금집행 균형 (예산=기투입+필수+분양불+준공후불)",
         f"='01_사업수지'!H{pl['집행검증']}", "불일치면 배분율 확인"),
        ("사업이익 부호", f"=IF('01_사업수지'!E{pl['사업이익']}>0,\"흑자\",\"적자\")",
         "적자면 재검토 판정 필요"),
        ("현금 부족 발생 여부", f"=IF('02_현금흐름'!C{cf['현금최저']}<0,\"자금부족 발생\",\"자금 충족\")",
         "부족 시 추가 조달 방안 필수"),
        ("PF 한도 내 조달 가능", f"='02_현금흐름'!C{cf['한도초과']}",
         "초과 시 한도 증액 또는 자기자본 추가 필요"),
        ("손익분기 분양률 < 100%",
         f"=IF('03_민감도'!B{sens['손익분기분양률']}<1,\"성립\",\"사업 성립 불가\")",
         "100% 이상이면 팔아도 원가 회수 불가"),
        ("수지표 수입 = 현금흐름 회수액",
         f"=IF(ROUND('02_현금흐름'!{cf['총액열']}{cf['수입계행']},0)"
         f">=ROUND('01_사업수지'!E{pl['분양총액']},0)-1,\"일치\","
         f"\"미회수 발생 — 분양률 누계 100% 미달\")",
         "사업기간 내 분양이 다 안 되면 수지표 수입은 과대"),
    ]
    for a, b, c in checks:
        ws2.cell(row=r, column=1, value=a)
        ws2.cell(row=r, column=2, value=b)
        ws2.cell(row=r, column=2).fill = FILL_OK
        ws2.cell(row=r, column=3, value=c).font = F_NORM
        for i in range(1, 4):
            ws2.cell(row=r, column=i).border = BORDER
        r += 1
    r += 1
    ws2.cell(row=r, column=1, value="※ 02_현금흐름 시트의 '검증' 열에 '불일치'가 있으면 배분 로직을 점검하십시오.")


# ---------------------------------------------------------------- 06_비례율
def sheet_jeongbi(wb, cfg, pl):
    j = cfg.get("jeongbi")
    if not j:
        return
    ws = wb.create_sheet("06_비례율")
    widths(ws, {"A": 30, "B": 20, "C": 50})
    ws["A1"] = "비례율·조합원 분담금 (가평가 — 법정 감정평가 아님)"
    ws["A1"].font = F_TITLE
    ws["A2"] = "※ 종전·종후자산은 감정평가법인등의 법정 평가액에 의합니다. 아래는 참고치입니다."
    r = 4
    r = head(ws, r, ["항목", "값", "산식·비고"])
    ws.cell(row=r, column=1, value="종후자산 총액 (총수입)")
    ws.cell(row=r, column=2, value=f"='01_사업수지'!E{pl['수입총계']}").number_format = NUM
    after = r
    r += 1
    # ★ 비례율 총사업비에서 제외할 항목은 '조합원이 현물출자한 종전자산'뿐이다.
    #   현금청산금·국공유지 매입비·주거이전비는 실제 현금지출이므로 반드시 포함한다.
    #   → 그룹 단위(토지비 전체) 제외 금지. 항목에 "비례율제외": true 를 붙인 행만 뺀다.
    excl_rows = pl.get("비례율제외행", [])
    if excl_rows:
        minus = "-" + "-".join(f"'01_사업수지'!H{x}" for x in excl_rows)
        names = ", ".join(pl.get("비례율제외명", []))
        ws.cell(row=r, column=1, value="총사업비 (현물출자분 제외)")
        ws.cell(row=r, column=2,
                value=f"='01_사업수지'!H{pl['지출총계']}{minus}").number_format = NUM
        ws.cell(row=r, column=3, value=f"제외 항목: {names} — 조합원 현물출자분 이중계상 방지")
    else:
        ws.cell(row=r, column=1, value="총사업비")
        ws.cell(row=r, column=2, value=f"='01_사업수지'!H{pl['지출총계']}").number_format = NUM
        ws.cell(row=r, column=3,
                value="제외 항목 없음. 조합원 현물출자 토지매입비를 계상했다면 "
                      "해당 항목에 \"비례율제외\": true 를 붙일 것")
    cost = r
    r += 1
    ws.cell(row=r, column=1, value="종전자산 총액")
    ws.cell(row=r, column=2, value=j.get("종전자산_천원", 0)).number_format = NUM
    ws.cell(row=r, column=2).fill = FILL_IN
    before = r
    r += 1
    ws.cell(row=r, column=1, value="★ 비례율").font = F_SUB
    ws.cell(row=r, column=2, value=f"=IFERROR((B{after}-B{cost})/B{before},0)").number_format = PCT
    ws.cell(row=r, column=3, value="(종후 − 총사업비) ÷ 종전")
    prop = r
    r += 1
    ws.cell(row=r, column=1, value="조합원 수")
    ws.cell(row=r, column=2, value=j.get("조합원수", 0)).number_format = NUM
    ws.cell(row=r, column=2).fill = FILL_IN
    nmem = r
    r += 1
    ws.cell(row=r, column=1, value="조합원 1인당 평균 종전자산")
    ws.cell(row=r, column=2, value=f"=IFERROR(B{before}/B{nmem},0)").number_format = NUM
    avg_before = r
    r += 1
    ws.cell(row=r, column=1, value="평균 권리가액")
    ws.cell(row=r, column=2, value=f"=B{avg_before}*B{prop}").number_format = NUM
    rights = r
    r += 1
    ws.cell(row=r, column=1, value="조합원 분양가 (평균)")
    ws.cell(row=r, column=2, value=j.get("조합원분양가_천원", 0)).number_format = NUM
    ws.cell(row=r, column=2).fill = FILL_IN
    memprice = r
    r += 1
    ws.cell(row=r, column=1, value="★ 평균 분담금").font = F_SUB
    ws.cell(row=r, column=2, value=f"=B{memprice}-B{rights}").number_format = NUM
    ws.cell(row=r, column=3, value="음수면 환급금")
    r += 2

    r = section(ws, r, "■ 공사비 민감도 → 비례율·분담금", 5)
    ws.cell(row=r, column=1, value="공사비 변동").font = F_SUB
    for j2, d in enumerate([-0.10, 0.0, 0.10, 0.20]):
        ws.cell(row=r, column=2 + j2, value=f"{d:+.0%}").font = F_SUB
    r += 1
    _b = pl.get('소계_건축비')
    bld = f"'01_사업수지'!$H${_b}" if _b else "0"
    ws.cell(row=r, column=1, value="비례율")
    for j2, d in enumerate([-0.10, 0.0, 0.10, 0.20]):
        ws.cell(row=r, column=2 + j2,
                value=f"=IFERROR((B{after}-(B{cost}+{bld}*{d}))/B{before},0)").number_format = PCT
    r += 1
    ws.cell(row=r, column=1, value="평균 분담금")
    for j2, d in enumerate([-0.10, 0.0, 0.10, 0.20]):
        L = get_column_letter(2 + j2)
        ws.cell(row=r, column=2 + j2,
                value=f"=B{memprice}-B{avg_before}*{L}{r-1}").number_format = NUM
    r += 2
    ws.cell(row=r, column=1,
            value="※ 조합원 개인정보(실명·주소·연락처)는 어떤 경우에도 기재하지 마십시오.").font = F_NORM


# ---------------------------------------------------------------- main
def ws_pl_max(wb):
    return wb["01_사업수지"].max_row


def main():
    # 플래그를 먼저 걷어낸 뒤 위치 인자를 파싱한다.
    # (플래그를 2번째 인자로 줘도 출력 파일명으로 오인하지 않도록)
    argv = sys.argv[1:]
    allow_warn = "--allow-warnings" in argv
    argv = [a for a in argv if not a.startswith("--")]
    if len(argv) < 2:
        print("Usage: python build_xlsx.py <config.json> <output.xlsx> [--allow-warnings]",
              file=sys.stderr)
        sys.exit(1)
    cfg_path, out_path = argv[0], argv[1]
    if not out_path.lower().endswith((".xlsx", ".xlsm")):
        print(f"[ERROR] 출력 파일은 .xlsx 여야 합니다: {out_path}", file=sys.stderr)
        sys.exit(1)
    with open(cfg_path, "r", encoding="utf-8") as f:
        cfg = json.load(f)

    # ---- 폐기된 키 감지 (Critical 회귀 방지)
    #      '토지비제외'는 그룹 단위 제외 방식의 잔재로, 비례율을 과대 산출시켰던 원인이다.
    #      항목 단위 '비례율제외'로 대체되었으므로 사용 시 명시적으로 알린다.
    if isinstance(cfg.get("jeongbi"), dict) and "토지비제외" in cfg["jeongbi"]:
        warn("jeongbi.토지비제외 는 폐기된 키입니다(무시됨). "
             "조합원 현물출자 항목에 \"비례율제외\": true 를 붙이십시오.")

    # ---- config 스키마 사전 검증 (raw traceback 방지)
    missing = [k for k in ("project", "site", "period", "revenue", "cost") if k not in cfg]
    if missing:
        print(f"[ERROR] config에 필수 블록이 없습니다: {', '.join(missing)}", file=sys.stderr)
        print("        config.sample.json 을 참고해 채우십시오.", file=sys.stderr)
        sys.exit(1)
    if not isinstance(cfg.get("revenue"), list) or not cfg["revenue"]:
        print("[ERROR] revenue 배열이 비어 있습니다. 최소 1개 항목이 필요합니다.", file=sys.stderr)
        sys.exit(1)
    if "name" not in cfg["project"]:
        print("[ERROR] project.name 이 없습니다.", file=sys.stderr)
        sys.exit(1)
    for k in ("인허가_개월", "공사_개월"):
        if k not in cfg["period"]:
            warn(f"period.{k} 가 없어 0으로 처리합니다.")

    wb = Workbook()
    wb.remove(wb.active)

    ref = sheet_input(wb, cfg)
    pl = sheet_pl(wb, cfg, ref)

    # 지출 소계 행 인덱스 확보 (sheet_pl 내부에서 기록되지 않은 경우 대비)
    ws_pl = wb["01_사업수지"]
    for row in range(1, ws_pl.max_row + 1):
        v = ws_pl.cell(row=row, column=2).value
        if isinstance(v, str) and v.endswith(" 소계"):
            pl["소계_" + v.replace(" 소계", "")] = row
    for g in ["토지비", "건축비", "부대비용", "금융비용"]:
        if "소계_" + g not in pl:
            warn(f"cost.{g} 블록이 없습니다. 현금흐름·민감도·비례율에서 해당 그룹은 0으로 처리합니다.")
            pl["소계_" + g] = None

    cf = sheet_cf(wb, cfg, ref, pl)

    # PF이자 행을 현금흐름 계산값에 연결 (실무 모델과 동일한 역참조).
    # 02_현금흐름의 지출에는 금융비용이 포함되지 않으므로 순환참조가 발생하지 않는다.
    linked = 0
    for row in range(1, ws_pl_max(wb) + 1):
        v = wb["01_사업수지"].cell(row=row, column=2).value
        if isinstance(v, str) and "PF이자" in v.replace(" ", ""):
            wb["01_사업수지"].cell(row=row, column=8).value = f"='02_현금흐름'!C{cf['이자합계']}"
            wb["01_사업수지"].cell(row=row, column=3).value = "02_현금흐름 이자 합계 (자동 연결)"
            linked += 1
    if linked == 0:
        warn("금융비용에 'PF이자' 항목이 없어 이자 자동연결을 하지 못했습니다. "
             "금융비용에 항목명 'PF이자'를 추가하십시오.")

    sens = sheet_sens(wb, cfg, pl)
    sheet_summary(wb, cfg, pl, cf, ref, sens)
    sheet_jeongbi(wb, cfg, pl)

    wb.save(out_path)

    # ---- 자체 검증: 수식이 가리키는 시트·행이 실제로 존재하는지 직접 대조한다.
    #      (openpyxl은 수식을 평가하지 않으므로 #REF! 문자열 탐지만으로는 검증이 되지 않는다)
    import re
    from openpyxl import load_workbook
    chk = load_workbook(out_path)
    maxrow = {ws.title: ws.max_row for ws in chk.worksheets}
    pat_cross = re.compile(r"'([^']+)'!\$?([A-Z]{1,3})\$?(\d+)")
    pat_lit = re.compile(r"#(REF|NAME|VALUE|DIV/0|N/A)")
    bad = 0
    for ws in chk.worksheets:
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if not isinstance(v, str) or not v.startswith("="):
                    if isinstance(v, str) and pat_lit.search(v):
                        print(f"[BROKEN] {ws.title}!{c.coordinate} 리터럴 오류: {v}",
                              file=sys.stderr)
                        bad += 1
                    continue
                if pat_lit.search(v):
                    print(f"[BROKEN] {ws.title}!{c.coordinate} = {v}", file=sys.stderr)
                    bad += 1
                    continue
                if "None" in v:
                    print(f"[BROKEN] {ws.title}!{c.coordinate} 미해결 참조: {v}", file=sys.stderr)
                    bad += 1
                    continue
                for sheet, col, rownum in pat_cross.findall(v):
                    if sheet not in maxrow:
                        print(f"[BROKEN] {ws.title}!{c.coordinate} → 없는 시트 '{sheet}'",
                              file=sys.stderr)
                        bad += 1
                    elif int(rownum) > maxrow[sheet]:
                        print(f"[BROKEN] {ws.title}!{c.coordinate} → '{sheet}'!{col}{rownum} "
                              f"는 사용 범위(max_row={maxrow[sheet]}) 밖", file=sys.stderr)
                        bad += 1
    print(f"OK: {out_path}")
    print(f"sheets={len(chk.sheetnames)} : {', '.join(chk.sheetnames)}")
    print(f"broken_refs={bad}")
    print(f"warnings={len(WARNINGS)}")
    if bad:
        sys.exit(2)
    if WARNINGS and not allow_warn:
        print("[FAIL] 경고가 있어 실패로 처리합니다. config를 수정하거나, "
              "의도한 것이라면 --allow-warnings 를 붙이고 보고서에 사유를 명시하십시오.",
              file=sys.stderr)
        sys.exit(3)


if __name__ == "__main__":
    main()
